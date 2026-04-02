import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
output = r'C:\Users\user\Downloads\FY2602_결산_GSheet용_v2.xlsx'

# RAW sheets: read values (data_only), trim 생산량_RAW and 매출채권_RAW
# INPUT/OUTPUT sheets: read formulas (preserve)

RAW_SHEETS = {'SAP_재무제표', '수주_RAW', '생산량_RAW', '매출상세_RAW', '매출채권_RAW', '인원현황_RAW', '2025_RAW'}

print("Loading workbooks...")
wb_val = load_workbook(file, data_only=True)   # for RAW sheets (values)
wb_form = load_workbook(file, data_only=False) # for INPUT/OUTPUT sheets (formulas)

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)  # remove default sheet

for sname in wb_form.sheetnames:
    print(f"Processing: {sname}...", end=' ')

    if sname in RAW_SHEETS:
        ws_src = wb_val[sname]

        if sname == '생산량_RAW':
            # Only rows 1~50000, and only rows with actual data
            ws_new = wb_out.create_sheet(sname)
            max_row = 50000
            written = 0
            for r_idx, row in enumerate(ws_src.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
                # Skip empty rows (after header) - check col G(7) and O(15)
                if r_idx > 1:
                    if row[6] is None and row[7] is None and row[14] is None:
                        continue
                ws_new.append(list(row))
                written += 1
            print(f"-> {written} rows (values)")

        elif sname == '매출채권_RAW':
            # Drop empty columns
            ws_src_form = wb_form[sname]
            # Find non-empty column indices (check rows 2-50)
            non_empty_idx = []
            for col in range(1, ws_src.max_column + 1):
                has_data = False
                for r in range(2, min(51, ws_src.max_row + 1)):
                    v = ws_src.cell(r, col).value
                    if v is not None and str(v).strip() not in ('', '0'):
                        has_data = True
                        break
                if has_data:
                    non_empty_idx.append(col)

            ws_new = wb_out.create_sheet(sname)
            for r_idx, row in enumerate(ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, values_only=True), 1):
                ws_new.append([row[i-1] for i in non_empty_idx])
            print(f"-> {ws_src.max_row} rows x {len(non_empty_idx)} cols (values)")

        else:
            # Normal RAW: copy values
            ws_new = wb_out.create_sheet(sname)
            for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, values_only=True):
                ws_new.append(list(row))
            print(f"-> {ws_src.max_row} rows (values)")

    else:
        # INPUT/OUTPUT sheets: preserve formulas
        ws_src = wb_form[sname]
        ws_new = wb_out.create_sheet(sname)

        for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
            for cell in row:
                new_cell = ws_new.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    try:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.number_format = cell.number_format
                    except Exception:
                        pass
        print(f"-> {ws_src.max_row} rows (formulas preserved)")

wb_out.save(output)
print(f"\nDone! Output: {output}")

import os
size_mb = os.path.getsize(output) / (1024 * 1024)
print(f"File size: {size_mb:.1f} MB")
