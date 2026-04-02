import openpyxl
from openpyxl.utils import get_column_letter
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산_GSheet용_v2.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb['① PL 입력']

# Row locations
revenue_row = 5    # I. 매출액
gp_row = 19        # III. 매출총이익
op_row = 53        # V. 영업이익
sga_row_data = 20  # IV. 판매비와관리비
gpm_row = 81
opm_row = 82
sga_pct_row = 83

# Find all data columns (header row 4, non-empty except col A)
data_cols = []
for c in range(2, 40):
    v = ws.cell(4, c).value
    if v and str(v).strip():
        data_cols.append((c, str(v).strip()))
        print(f'  Col {get_column_letter(c)}: {v}')

print(f'\nTotal data columns: {len(data_cols)}')

# Fix GPM, OPM, SGA% formulas for all data columns
fixed = 0
for col_idx, col_name in data_cols:
    col = get_column_letter(col_idx)
    rev_cell = f'{col}{revenue_row}'
    gp_cell = f'{col}{gp_row}'
    op_cell = f'{col}{op_row}'
    sga_cell = f'{col}{sga_row_data}'

    # GPM = 매출총이익 / 매출액
    gpm_formula = f'=IF({rev_cell}<>0,{gp_cell}/{rev_cell},"")'
    ws.cell(gpm_row, col_idx).value = gpm_formula

    # OPM = 영업이익 / 매출액
    opm_formula = f'=IF({rev_cell}<>0,{op_cell}/{rev_cell},"")'
    ws.cell(opm_row, col_idx).value = opm_formula

    # SGA% = 판관비 / 매출액
    sga_formula = f'=IF({rev_cell}<>0,{sga_cell}/{rev_cell},"")'
    ws.cell(sga_pct_row, col_idx).value = sga_formula

    fixed += 1

# Also fix AA-AE SUM formulas (FY25 quarterly aggregates if they exist)
# These reference O-Z which are 26-01 to 26-12 monthly columns
# Let's verify what O-Z map to and set correct quarterly aggregates
# Find column letters for 26-01..26-12
cols_26 = [(c, h) for c, h in data_cols if h.startswith('26-')]
print(f'\nFY26 monthly columns: {[(get_column_letter(c), h) for c, h in cols_26]}')

# Fix AA-AE if they're quarterly sums over FY26 months
if len(cols_26) == 12:
    col_letters_26 = [get_column_letter(c) for c, _ in cols_26]
    q1 = col_letters_26[0:3]
    q2 = col_letters_26[3:6]
    q3 = col_letters_26[6:9]
    q4 = col_letters_26[9:12]

    for pct_row in [gpm_row, opm_row, sga_pct_row]:
        for aa_col, q_cols in zip([27, 28, 29, 30], [q1, q2, q3, q4]):
            ws.cell(pct_row, aa_col).value = f'=AVERAGE({q_cols[0]}{revenue_row}:{q_cols[2]}{pct_row})'

wb.save(file)
print(f'\nFixed {fixed} columns in GPM/OPM/SGA% rows.')
print(f'Saved: {file}')
