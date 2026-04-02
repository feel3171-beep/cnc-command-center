"""NAS file reading tools — 경영팀(X:), 인사팀(Y:), 경영정보팀(Z:) NAS 접근"""

import json
import os
from pathlib import Path

# NAS drive mappings
NAS_DRIVES = {
    "경영팀": "X:/",
    "인사팀": "Y:/",
    "경영정보팀": "Z:/",
}

TOOL_LIST_NAS_FILES = {
    "name": "list_nas_files",
    "description": "NAS 드라이브의 파일/폴더 목록을 조회합니다. 경영팀(X:), 인사팀(Y:), 경영정보팀(Z:) 접근 가능.",
    "input_schema": {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "조회할 경로. 예: 'X:/02 자금/1. 자금일보', 'Y:/Recruit', 'Z:/00.경영정보팀_경영계획'",
            },
            "pattern": {
                "type": "string",
                "description": "파일명 필터 (선택). 예: '*.xlsx', '*채용*', '*2026*'",
            },
        },
        "required": ["path"],
    },
}

TOOL_READ_NAS_EXCEL = {
    "name": "read_nas_excel",
    "description": "NAS의 Excel 파일을 읽어 데이터를 반환합니다. 시트명, 헤더, 데이터를 포함합니다.",
    "input_schema": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Excel 파일 전체 경로. 예: 'Y:/Recruit/Total_채용진행사항.xlsx'",
            },
            "sheet_name": {
                "type": "string",
                "description": "시트명 (선택). 생략 시 첫 번째 시트.",
            },
            "max_rows": {
                "type": "integer",
                "description": "최대 행 수 (기본 100). 큰 파일은 제한 필요.",
            },
        },
        "required": ["file_path"],
    },
}


def handle_list_nas_files(tool_input: dict) -> str:
    """List files in NAS directory."""
    path = tool_input["path"]
    pattern = tool_input.get("pattern", "")

    try:
        p = Path(path)
        if not p.exists():
            return json.dumps({"error": f"경로 없음: {path}"}, ensure_ascii=False)

        items = []
        for item in sorted(p.iterdir()):
            if item.name.startswith("~$") or item.name.startswith("."):
                continue
            if pattern and not __import__("fnmatch").fnmatch(item.name, pattern):
                continue
            items.append({
                "name": item.name,
                "type": "dir" if item.is_dir() else "file",
                "size_kb": round(item.stat().st_size / 1024) if item.is_file() else None,
                "modified": __import__("datetime").datetime.fromtimestamp(
                    item.stat().st_mtime
                ).strftime("%Y-%m-%d") if item.is_file() else None,
            })

        if len(items) > 50:
            items = items[:50]
            return json.dumps({"items": items, "truncated": True, "total": len(list(p.iterdir()))}, ensure_ascii=False)

        return json.dumps({"items": items, "count": len(items)}, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


def handle_read_nas_excel(tool_input: dict) -> str:
    """Read Excel file from NAS."""
    file_path = tool_input["file_path"]
    sheet_name = tool_input.get("sheet_name")
    max_rows = tool_input.get("max_rows", 100)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb[sheet_names[0]]

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])

        wb.close()

        return json.dumps({
            "file": file_path,
            "sheet": ws.title,
            "all_sheets": sheet_names,
            "rows": len(rows),
            "data": rows,
        }, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": str(e), "file": file_path}, ensure_ascii=False)
