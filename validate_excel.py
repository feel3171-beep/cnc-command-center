import openpyxl
import sys
import io
import json
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'

# Load with formulas
wb = openpyxl.load_workbook(file)
# Load with values
wb_val = openpyxl.load_workbook(file, data_only=True)

results = {}

for sname in wb.sheetnames:
    ws = wb[sname]
    ws_val = wb_val[sname]

    sheet_info = {
        'rows': ws.max_row,
        'cols': ws.max_column,
        'formula_count': 0,
        'empty_in_data_rows': 0,
        'data_types': defaultdict(int),
        'formula_errors': [],
        'formulas_sample': [],
        'merged_cells': len(ws.merged_cells.ranges),
        'sum_checks': [],
    }

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            if c.value is None:
                continue
            if c.data_type == 'f':
                sheet_info['formula_count'] += 1
                sheet_info['data_types']['formula'] += 1
                if len(sheet_info['formulas_sample']) < 5:
                    sheet_info['formulas_sample'].append(f'{c.coordinate}={c.value}')
            elif c.data_type == 'n':
                sheet_info['data_types']['number'] += 1
            elif c.data_type == 's':
                sheet_info['data_types']['string'] += 1
            elif c.data_type == 'b':
                sheet_info['data_types']['boolean'] += 1
            elif c.data_type == 'd':
                sheet_info['data_types']['date'] += 1
            else:
                sheet_info['data_types'][c.data_type] += 1

    # Check for error values in data_only workbook
    for row in ws_val.iter_rows(min_row=1, max_row=ws_val.max_row, max_col=ws_val.max_column):
        for c in row:
            if c.value is not None and isinstance(c.value, str):
                if c.value in ('#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!'):
                    sheet_info['formula_errors'].append(f'{c.coordinate}={c.value}')

    sheet_info['data_types'] = dict(sheet_info['data_types'])
    results[sname] = sheet_info

# Print summary
print("=" * 70)
print("정합성 검증 결과 요약")
print("=" * 70)

for sname, info in results.items():
    print(f"\n--- {sname} ---")
    print(f"  크기: {info['rows']}행 x {info['cols']}열")
    print(f"  데이터 타입: {info['data_types']}")
    print(f"  수식 수: {info['formula_count']}")
    print(f"  병합셀: {info['merged_cells']}")
    if info['formulas_sample']:
        print(f"  수식 샘플:")
        for f in info['formulas_sample']:
            print(f"    {f}")
    if info['formula_errors']:
        print(f"  ⚠ 수식 오류 ({len(info['formula_errors'])}건):")
        for e in info['formula_errors'][:20]:
            print(f"    {e}")
    else:
        print(f"  ✓ 수식 오류 없음")

# Cross-sheet validation: check INPUT sheets reference RAW sheets properly
print("\n" + "=" * 70)
print("시트간 참조 검증")
print("=" * 70)

for sname in wb.sheetnames:
    ws = wb[sname]
    refs = defaultdict(int)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            if c.data_type == 'f' and c.value:
                formula = str(c.value)
                for other_sheet in wb.sheetnames:
                    if other_sheet != sname and other_sheet in formula:
                        refs[other_sheet] += 1
    if refs:
        print(f"\n  {sname} → 참조하는 시트:")
        for ref_sheet, cnt in refs.items():
            print(f"    → {ref_sheet}: {cnt}건")
