import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb['매출채권_RAW']

print(f"max_row={ws.max_row}, max_col={ws.max_column}")

# Headers - check first row
print("\n=== 헤더 샘플 ===")
for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 50, 100, 200, 400, 600, 838]:
    c = ws.cell(1, col)
    cl = openpyxl.utils.get_column_letter(col)
    if c.value:
        print(f"  {cl}1 = {str(c.value)[:60]}")

# Find actual last non-empty column
print("\n=== 실제 데이터 폭 파악 ===")
last_col = 0
for col in range(ws.max_column, 0, -1):
    has_data = False
    for r in [1, 2, 5, 10, 100]:
        c = ws.cell(r, col)
        if c.value is not None and str(c.value).strip() != '':
            has_data = True
            break
    if has_data:
        last_col = col
        break
print(f"  마지막 실데이터 열: {last_col} ({openpyxl.utils.get_column_letter(last_col)})")

# Find actual last row
last_row = 0
for r in range(ws.max_row, max(1, ws.max_row - 1000), -1):
    for col in [1, 2, 6]:
        c = ws.cell(r, col)
        if c.value is not None and str(c.value).strip() != '':
            last_row = max(last_row, r)
            break
print(f"  마지막 실데이터 행: {last_row}")

# Sample data
print("\n=== 데이터 샘플 (Row 2) ===")
for col in range(1, min(last_col + 1, 20)):
    c = ws.cell(2, col)
    cl = openpyxl.utils.get_column_letter(col)
    print(f"  {cl}2 = {str(c.value)[:60]}")

# Also check 매출상세_RAW actual column usage
ws2 = wb['매출상세_RAW']
last_col2 = 0
for col in range(ws2.max_column, 0, -1):
    has_data = False
    for r in [1, 2, 5, 18, 100]:
        c = ws2.cell(r, col)
        if c.value is not None and str(c.value).strip() != '':
            has_data = True
            break
    if has_data:
        last_col2 = col
        break
print(f"\n매출상세_RAW 마지막 실데이터 열: {last_col2} ({openpyxl.utils.get_column_letter(last_col2)})")

# 수주_RAW actual column usage
ws3 = wb['수주_RAW']
last_col3 = 0
for col in range(ws3.max_column, 0, -1):
    has_data = False
    for r in [1, 2, 5, 100]:
        c = ws3.cell(r, col)
        if c.value is not None and str(c.value).strip() != '':
            has_data = True
            break
    if has_data:
        last_col3 = col
        break
print(f"수주_RAW 마지막 실데이터 열: {last_col3} ({openpyxl.utils.get_column_letter(last_col3)})")
