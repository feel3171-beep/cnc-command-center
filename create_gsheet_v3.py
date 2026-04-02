import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
output = r'C:\Users\user\Downloads\FY2602_결산_GSheet용_v2.xlsx'

RAW_SHEETS = {'SAP_재무제표', '수주_RAW', '생산량_RAW', '매출상세_RAW', '매출채권_RAW', '인원현황_RAW', '2025_RAW'}

# Pre-compute 매출채권_RAW non-empty columns using pandas (fast)
print("Finding non-empty columns in 매출채권_RAW using pandas...")
df_ar = pd.read_excel(file, sheet_name='매출채권_RAW', header=0, nrows=100)
ar_nonempty_cols = [i for i, c in enumerate(df_ar.columns) if df_ar.iloc[:, i].notna().any()]
print(f"  {len(ar_nonempty_cols)} cols with data (out of {len(df_ar.columns)})")

print("Loading workbooks...")
wb_val = load_workbook(file, data_only=True)
wb_form = load_workbook(file, data_only=False)

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

for sname in wb_form.sheetnames:
    print(f"Processing: {sname}...", end=' ', flush=True)

    if sname in RAW_SHEETS:
        ws_src = wb_val[sname]

        if sname == '생산량_RAW':
            ws_new = wb_out.create_sheet(sname)
            written = 0
            for r_idx, row in enumerate(ws_src.iter_rows(min_row=1, max_row=50000, values_only=True), 1):
                if r_idx > 1 and row[6] is None and row[7] is None and row[14] is None:
                    continue
                ws_new.append(list(row))
                written += 1
            print(f"-> {written} rows (values)")

        elif sname == '매출채권_RAW':
            ws_new = wb_out.create_sheet(sname)
            for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, values_only=True):
                ws_new.append([row[i] for i in ar_nonempty_cols])
            print(f"-> {ws_src.max_row} rows x {len(ar_nonempty_cols)} cols (values)")

        else:
            ws_new = wb_out.create_sheet(sname)
            for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, values_only=True):
                ws_new.append(list(row))
            print(f"-> {ws_src.max_row} rows (values)")

    else:
        # INPUT/OUTPUT: preserve formulas (skip style copy for speed)
        ws_src = wb_form[sname]
        ws_new = wb_out.create_sheet(sname)
        for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
            for cell in row:
                if cell.value is not None:
                    ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
        print(f"-> {ws_src.max_row} rows (formulas preserved)")

wb_out.save(output)
print(f"\nDone! Output: {output}")

import os
size_mb = os.path.getsize(output) / (1024 * 1024)
print(f"File size: {size_mb:.1f} MB")
