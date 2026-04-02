import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb['생산량_RAW']

print(f"max_row={ws.max_row}, max_col={ws.max_column}")

# Header
print("\n=== 헤더 (Row 1) ===")
for col in range(1, 25):
    c = ws.cell(1, col)
    if c.value is not None:
        print(f"  {c.coordinate} = {c.value}")

# Find actual last data row by checking column O (date column, seems to have actual data)
print("\n=== 실제 데이터 범위 파악 ===")
# Check a few key columns for actual data
for check_col in [7, 8, 15]:  # G, H, O
    col_letter = openpyxl.utils.get_column_letter(check_col)
    last = 0
    for r in range(2, min(ws.max_row + 1, 50000)):
        c = ws.cell(r, check_col)
        if c.value is not None and str(c.value).strip() != '' and str(c.value) != '0':
            last = r
    print(f"  Column {col_letter}: 마지막 실제 데이터 행 = {last}")

# Sample rows near the boundary
print("\n=== 데이터 경계 부근 샘플 ===")
for r in [2, 3, 100, 500, 1000, 2000, 3000, 4000, 5000]:
    vals = []
    for col in [1, 2, 4, 7, 8, 15]:
        c = ws.cell(r, col)
        v = str(c.value)[:30] if c.value is not None else 'None'
        vals.append(f"{openpyxl.utils.get_column_letter(col)}{r}={v}")
    print(f"  {' | '.join(vals)}")

# Check formulas in original (non-data_only) workbook for columns A-D
wb2 = openpyxl.load_workbook(file)
ws2 = wb2['생산량_RAW']
print("\n=== 수식 컬럼 (A~F) 확인 ===")
for col in range(1, 7):
    c = ws2.cell(2, col)
    cl = openpyxl.utils.get_column_letter(col)
    if c.value:
        print(f"  {cl}2 = {str(c.value)[:100]} (type={c.data_type})")
    # Check if formula extends to very far rows
    for test_row in [100, 1000, 10000, 100000, 500000, 1048576]:
        tc = ws2.cell(test_row, col)
        if tc.value is not None:
            print(f"    {cl}{test_row} = {str(tc.value)[:60]} (type={tc.data_type})")
        else:
            print(f"    {cl}{test_row} = None")
            break
