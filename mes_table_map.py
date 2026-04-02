from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── 색상/스타일 정의 ───
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill_rt = PatternFill('solid', fgColor='1F4E79')  # 실시간 - 진한파랑
header_fill_if = PatternFill('solid', fgColor='2E75B6')  # I/F - 중간파랑
header_fill_out = PatternFill('solid', fgColor='548235')  # OUT - 초록
header_fill_master = PatternFill('solid', fgColor='7F7F7F')  # 마스터 - 회색
header_fill_etc = PatternFill('solid', fgColor='BF8F00')  # 기타 - 황금

rt_fill = PatternFill('solid', fgColor='DAEEF3')
if_fill = PatternFill('solid', fgColor='D6E4F0')
out_fill = PatternFill('solid', fgColor='E2EFDA')
master_fill = PatternFill('solid', fgColor='EDEDED')
etc_fill = PatternFill('solid', fgColor='FFF2CC')

title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
section_font = Font(name='Arial', bold=True, size=12, color='1F4E79')
normal_font = Font(name='Arial', size=10)
num_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def style_header(ws, row, cols, fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if c == 3:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')

# ═══════════════════════════════════════════════
# 시트1: 전체 테이블 맵
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = '전체 테이블 맵'
ws1.sheet_properties.tabColor = '1F4E79'

ws1.merge_cells('A1:H1')
ws1['A1'] = 'MES DB 테이블 맵 분석 (2026-03-29 기준)'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:H2')
ws1['A2'] = 'Server: 192.161.0.16 / Database: MES / 데이터 있는 테이블만 정리 (백업 테이블 제외)'
ws1['A2'].font = Font(name='Arial', size=9, color='808080')

headers = ['구분', '테이블명', '건수', '최종수정일', '모듈', '테이블유형', '설명', '실시간 활용']
r = 4

# ── 실시간 트랜잭션 ──
ws1.merge_cells(f'A{r}:H{r}')
ws1[f'A{r}'] = '■ 실시간 트랜잭션 (3월 현재 계속 업데이트 중 - 대시보드 핵심)'
ws1[f'A{r}'].font = section_font
ws1[f'A{r}'].fill = PatternFill('solid', fgColor='DAEEF3')
r += 1

for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8, header_fill_rt)
r += 1

rt_data = [
    ['RT-01','MWIPLOTHIS',2172402,'실시간','MES코어','LOT이력(트랜잭션)','생산 LOT별 CV/IN/OUT 트랜잭션','★ 생산실적 실시간 모니터링'],
    ['RT-02','MWIPLOTSTS',369870,'03-18','MES코어','LOT현황(상태)','LOT 현재 상태/위치/수량','★ LOT 진행상태 추적'],
    ['RT-03','MWIPORDSTS',14961,'03-26','MES코어','작업지시(상태)','생산 작업지시 현황','★ 지시 대비 실적 추적'],
    ['RT-04','MWIPNWKSTS',5522,'실시간','MES코어','비가동(상태)','라인별 비가동 시간/유형/사유','★ 가동률/비가동 분석'],
    ['RT-05','MWIPLOTCQH',350151,'03-25','MES코어','품질확인이력','LOT별 품질확인 이력','품질 합부 현황'],
    ['RT-06','MINVLOTHIS',5110691,'03-16','MES코어','자재LOT이력','자재 입출고 트랜잭션','자재 소진/입고 추적'],
    ['RT-07','MINVLOTISS',2348210,'03-21','MES코어','자재출고','원부자재 출고 이력','원부자재 소진 현황'],
    ['RT-08','MINVLOTSTS',746714,'03-12','MES코어','자재LOT현황','자재 LOT 현재 상태','재고 현재 상태'],
    ['RT-09','MQCMREQSTS',113377,'03-20','MES코어','품질검사요청','품질검사 요청 현황','검사 진행/대기 현황'],
    ['RT-10','CQCMAPRSTS',34112,'03-24','커스텀','품질승인','품질 승인 현황','승인 대기/완료'],
    ['RT-11','MWIPBOMCMP',788643,'03-24','MES코어','BOM자재','BOM 자재 구성','자재 소요량 계산'],
]
for row_data in rt_data:
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 8, rt_fill)
    r += 1

# ── I/F 테이블 ──
r += 1
ws1.merge_cells(f'A{r}:H{r}')
ws1[f'A{r}'] = '■ I/F 인터페이스 (ERP ↔ MES 동기화, 주기적 배치)'
ws1[f'A{r}'].font = section_font
ws1[f'A{r}'].fill = PatternFill('solid', fgColor='D6E4F0')
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8, header_fill_if)
r += 1

if_data = [
    ['IF-01','IWIPMATDEF',1180320,'03-12','I/F','자재마스터','ERP→MES 자재 마스터 동기화','품목코드/명 참조'],
    ['IF-02','IWIPBOMDEF',852619,'02-04','I/F','BOM정의','ERP→MES BOM 동기화','BOM 기준정보'],
    ['IF-03','IWIPBOMCMP',725819,'(배치)','I/F','BOM자재','ERP→MES BOM 자재 동기화',''],
    ['IF-04','IWIPFRMOPR',442030,'(배치)','I/F','공정프레임','ERP→MES 공정 동기화',''],
    ['IF-05','IWIPORDSTS',91859,'02-11','I/F','작업지시','ERP→MES 작업지시 인터페이스','작업지시 원본 추적'],
    ['IF-06','IINVDLVDTL',23903,'01-19','I/F','납품상세','ERP→MES 납품 데이터','납품 일정 추적'],
    ['IF-07','IINVDLVMST',11545,'11-18','I/F','납품마스터','납품 마스터 정보',''],
    ['IF-08','IINVSHPDTL',11559,'12-18','I/F','출하상세','출하 상세 정보','출하 현황'],
    ['IF-09','IQCMINCSTS',15610,'12-21','I/F','수입검사','ERP→MES 수입검사','입고 품질'],
    ['IF-10','IINVREQDTL',16445,'(배치)','I/F','자재요청상세','자재 요청 상세',''],
    ['IF-11','IINVREQMST',15886,'(배치)','I/F','자재요청마스터','자재 요청 마스터',''],
]
for row_data in if_data:
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 8, if_fill)
    r += 1

# ── 아웃바운드 ──
r += 1
ws1.merge_cells(f'A{r}:H{r}')
ws1[f'A{r}'] = '■ O(아웃바운드) - MES → 외부 시스템 전송'
ws1[f'A{r}'].font = section_font
ws1[f'A{r}'].fill = PatternFill('solid', fgColor='E2EFDA')
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8, header_fill_out)
r += 1

out_data = [
    ['OUT-01','OINVWIPLOT',4106930,'12-17','아웃바운드','LOT전송','LOT 데이터 외부 전송',''],
    ['OUT-02','OWIPLOTSTS',354012,'03-26','아웃바운드','LOT상태전송','LOT 상태 전송',''],
    ['OUT-03','OWIPORDSTS',44827,'11-05','아웃바운드','작업지시전송','작업지시 결과 전송',''],
    ['OUT-04','OINVLOTISS',1903797,'(배치)','아웃바운드','출고전송','출고 데이터 전송',''],
    ['OUT-05','OINVLOTMVH',837165,'(배치)','아웃바운드','이동전송','LOT 이동 전송',''],
]
for row_data in out_data:
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 8, out_fill)
    r += 1

# ── 마스터 ──
r += 1
ws1.merge_cells(f'A{r}:H{r}')
ws1[f'A{r}'] = '■ 마스터/정의 테이블 (거의 변경 없음, 조인용)'
ws1[f'A{r}'].font = section_font
ws1[f'A{r}'].fill = PatternFill('solid', fgColor='EDEDED')
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8, header_fill_master)
r += 1

master_data = [
    ['MST-01','MWIPMATDEF',424104,'03-21','MES코어','자재/제품 마스터','품목코드→품목명/규격/단위 매핑','모든 조회의 기본 조인'],
    ['MST-02','MMDMMATDEF',424227,'(마스터)','MES코어','MDM 자재 마스터','MDM 기준 자재 정보',''],
    ['MST-03','MINVLBLDEF',873880,'03-03','MES코어','라벨 정의','라벨 마스터 정보',''],
    ['MST-04','MWIPLINDEF',268,'(마스터)','MES코어','라인 정의','라인코드→라인명/유형/교대시간','★ 라인유형(충진/타정/포장) 구분'],
    ['MST-05','MWIPOPRDEF',651,'(마스터)','MES코어','공정 정의','공정코드→공정명 매핑','공정별 분석'],
    ['MST-06','MWIPFLWDEF',11,'(마스터)','MES코어','플로우 정의','생산 플로우 정의',''],
    ['MST-07','MWIPCUSDEF',2433,'(마스터)','MES코어','거래처 정의','거래처코드→거래처명','거래처별 분석'],
    ['MST-08','MWIPVENDEF',4371,'(마스터)','MES코어','공급업체 정의','공급업체 마스터',''],
    ['MST-09','MWIPBOMSET',102607,'02-02','MES코어','BOM 세트','BOM 세트 정의',''],
    ['MST-10','MWIPBOMOPR',137883,'(마스터)','MES코어','BOM 공정','BOM별 공정 순서',''],
    ['MST-11','MADMTBLDAT',11114,'(마스터)','MES코어','코드 테이블','공통코드 값 정의','코드값 해석'],
    ['MST-12','MADMUSRDEF',727,'(마스터)','MES코어','사용자 정의','사용자 마스터',''],
    ['MST-13','MWIPSTOLOC',7366,'(마스터)','MES코어','저장위치','창고/위치 정의',''],
]
for row_data in master_data:
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 8, master_fill)
    r += 1

# ── 마감/집계/리포트 ──
r += 1
ws1.merge_cells(f'A{r}:H{r}')
ws1[f'A{r}'] = '■ 마감/집계/리포트 (일/월 마감 시 생성)'
ws1[f'A{r}'].font = section_font
ws1[f'A{r}'].fill = PatternFill('solid', fgColor='FFF2CC')
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8, header_fill_etc)
r += 1

etc_data = [
    ['RPT-01','RWIPLOTFIN',31267,'(마감)','리포트','LOT완료 리포트','완료 LOT 집계','기간별 실적 리포트'],
    ['RPT-02','RWIPLOTFOU',19176,'(마감)','리포트','LOT생산 리포트','생산 LOT 집계',''],
    ['RPT-03','RSUMWIPMOV',4742,'(마감)','리포트','생산이동 집계','생산 이동 집계',''],
    ['RPT-04','MINVDLVMST',9551,'02-21','MES코어','납품 마스터','납품 마스터 정보','납품 현황 (마감성)'],
    ['RPT-05','MINVDLVDTL',20152,'(마감)','MES코어','납품 상세','납품 상세 데이터',''],
    ['RPT-06','MINVSHPMST',2298,'(마감)','MES코어','출하 마스터','출하 마스터','출하 현황 (마감성)'],
    ['RPT-07','MINVSHPDTL',9210,'(마감)','MES코어','출하 상세','출하 상세 데이터',''],
    ['RPT-08','MINVREQMST',14082,'(마감)','MES코어','자재요청 마스터','자재 요청 마스터',''],
    ['RPT-09','MINVREQDTL',127058,'(마감)','MES코어','자재요청 상세','자재 요청 상세',''],
    ['RPT-10','MQCMLOTDAT',898443,'(마감)','MES코어','품질LOT데이터','LOT별 품질 데이터',''],
]
for row_data in etc_data:
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 8, etc_fill)
    r += 1

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 35
ws1.column_dimensions['H'].width = 30

# ═══════════════════════════════════════════════
# 시트2: 실시간 대시보드 설계
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('실시간 대시보드 설계')
ws2.sheet_properties.tabColor = '2E75B6'

ws2.merge_cells('A1:F1')
ws2['A1'] = '실시간 대시보드 설계 - 테이블 조인 관계 & 분석 시나리오'
ws2['A1'].font = title_font
ws2.row_dimensions[1].height = 30

# 핵심 데이터 흐름
r = 3
ws2.merge_cells(f'A{r}:F{r}')
ws2[f'A{r}'] = '1. 핵심 데이터 흐름 (ERP → MES → 분석)'
ws2[f'A{r}'].font = section_font
r += 1

flow_headers = ['단계', '테이블', 'JOIN KEY', '역할', '업데이트 주기', '비고']
for c, h in enumerate(flow_headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 6, header_fill_rt)
r += 1

flows = [
    ['① 작업지시 수신','IWIPORDSTS','FACTORY_CODE + ORDER_NO','ERP에서 작업지시 수신','배치 (일 1~2회)','ERP 원본 데이터'],
    ['② 작업지시 실행','MWIPORDSTS','FACTORY_CODE + ORDER_NO','MES에서 작업지시 실행/관리','실시간','ORD_STATUS로 진행상태'],
    ['③ LOT 생산 트랜잭션','MWIPLOTHIS','ORDER_NO + LOT_NO','CV/IN/OUT 등 LOT별 이벤트','실시간 (초단위)','TRAN_CODE=CV가 생산완료'],
    ['④ LOT 현재 상태','MWIPLOTSTS','FACTORY_CODE + LOT_NO','LOT의 현재 위치/상태/수량','실시간','현재 스냅샷'],
    ['⑤ 품질 확인','MWIPLOTCQH','LOT_NO','LOT별 품질확인 이력','실시간','합격/불합격/보류'],
    ['⑥ 비가동 등록','MWIPNWKSTS','FACTORY_CODE + LINE_CODE','비가동 시간/사유 기록','실시간','가동률 계산의 핵심'],
]
for fd in flows:
    for c, v in enumerate(fd, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 6, rt_fill)
    r += 1

# 분석 시나리오
r += 2
ws2.merge_cells(f'A{r}:F{r}')
ws2[f'A{r}'] = '2. 실시간 분석 시나리오'
ws2[f'A{r}'].font = section_font
r += 1

scenario_headers = ['시나리오', '주요 테이블', '보조 테이블(조인)', '핵심 지표', '갱신 주기', '난이도']
for c, h in enumerate(scenario_headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 6, header_fill_if)
r += 1

scenarios = [
    ['라인별 시간당 생산량','MWIPLOTHIS (CV)','MWIPLINDEF, MWIPMATDEF','시간당 생산수량, 라인가동률','1분','★'],
    ['작업지시 진척률','MWIPORDSTS','MWIPLOTHIS, MWIPMATDEF','지시수량 vs 생산수량 비율','5분','★'],
    ['비가동 분석','MWIPNWKSTS','MWIPLINDEF, MADMTBLDAT','비가동 시간/횟수/사유별 비중','5분','★★'],
    ['LOT 추적 (어디까지 왔나)','MWIPLOTSTS','MWIPORDSTS, MWIPLINDEF','LOT 현재 위치/상태','실시간','★'],
    ['품질 합격률','MWIPLOTCQH + MQCMREQSTS','MWIPMATDEF','품목별/라인별 합격률','10분','★★'],
    ['자재 소진 현황','MINVLOTISS + MINVLOTSTS','MWIPMATDEF','원자재 재고 vs 소요량','10분','★★'],
    ['납품 대비 생산 진척','MWIPORDSTS + IINVDLVDTL','MWIPCUSDEF','납기일 대비 생산 완료율','30분','★★★'],
    ['공장간 생산성 비교','MWIPLOTHIS','MWIPLINDEF, MWIPFACDEF','공장별 동일 제품 생산성','1시간','★★'],
]
for sd in scenarios:
    for c, v in enumerate(sd, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 6)
    r += 1

# 조인 관계도
r += 2
ws2.merge_cells(f'A{r}:F{r}')
ws2[f'A{r}'] = '3. 주요 JOIN KEY 정리'
ws2[f'A{r}'].font = section_font
r += 1

join_headers = ['FROM 테이블', 'TO 테이블', 'JOIN KEY', '관계', '용도', '']
for c, h in enumerate(join_headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 6, header_fill_master)
r += 1

joins = [
    ['MWIPORDSTS','MWIPLOTHIS','ORDER_NO','1:N','작업지시 → LOT 이력',''],
    ['MWIPLOTHIS','MWIPLOTSTS','FACTORY_CODE + LOT_NO','N:1','이력 → 현재상태',''],
    ['MWIPLOTHIS','MWIPLINDEF','FACTORY_CODE + LINE_CODE','N:1','LOT → 라인정보',''],
    ['MWIPLOTHIS','MWIPMATDEF','FACTORY_CODE + MAT_CODE','N:1','LOT → 품목정보',''],
    ['MWIPORDSTS','MWIPLINDEF','FACTORY_CODE + LINE_CODE','N:1','지시 → 라인정보',''],
    ['MWIPORDSTS','MWIPMATDEF','FACTORY_CODE + MAT_CODE','N:1','지시 → 품목정보',''],
    ['MWIPNWKSTS','MWIPLINDEF','FACTORY_CODE + LINE_CODE','N:1','비가동 → 라인정보',''],
    ['MWIPLOTCQH','MWIPLOTSTS','FACTORY_CODE + LOT_NO','N:1','품질 → LOT상태',''],
    ['IWIPORDSTS','MWIPORDSTS','FACTORY_CODE + ORDER_NO','1:1','ERP지시 → MES지시',''],
    ['MWIPORDSTS','MWIPCUSDEF','FACTORY_CODE + CUSTOMER_CODE','N:1','지시 → 거래처',''],
]
for jd in joins:
    for c, v in enumerate(jd, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 6)
    r += 1

for c in range(1, 7):
    ws2.column_dimensions[get_column_letter(c)].width = [25, 22, 30, 22, 30, 10][c-1]

# ═══════════════════════════════════════════════
# 시트3: 확인 필요 코드값
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('코드값 확인 필요')
ws3.sheet_properties.tabColor = 'BF8F00'

ws3.merge_cells('A1:F1')
ws3['A1'] = '코드값 확인 필요 목록 - 실시간 분석을 위해 의미 파악 필요'
ws3['A1'].font = title_font
ws3.row_dimensions[1].height = 30

r = 3
code_headers = ['우선순위', '테이블', '컬럼', '예상 의미', '확인 방법', '분석 활용']
for c, h in enumerate(code_headers, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 6, header_fill_rt)
r += 1

codes = [
    ['★★★','MWIPORDSTS','ORD_STATUS','작업지시 상태 (대기/진행/완료/마감)','SELECT DISTINCT ORD_STATUS','지시 진척률 계산'],
    ['★★★','MWIPLOTHIS','TRAN_CODE','트랜잭션 유형 (CV/IN/OUT 등)','SELECT DISTINCT TRAN_CODE','CV=생산완료 확인'],
    ['★★★','MWIPLINDEF','LINE_TYPE','라인유형 (충진/타정/포장/제조 등)','SELECT DISTINCT LINE_TYPE','라인유형별 분석'],
    ['★★★','MWIPORDSTS','ORD_TYPE','작업지시 유형','SELECT DISTINCT ORD_TYPE','지시 유형별 필터'],
    ['★★','MWIPLOTSTS','LOT_STATUS','LOT 상태','SELECT DISTINCT LOT_STATUS','LOT 진행상태'],
    ['★★','MWIPNWKSTS','NONWORK_TYPE','비가동 유형 (계획/비계획)','SELECT DISTINCT NONWORK_TYPE','비가동 분류'],
    ['★★','MWIPNWKSTS','NONWORK_CODE','비가동 사유코드','SELECT DISTINCT NONWORK_CODE','비가동 사유별 분석'],
    ['★★','MWIPNWKSTS','NONWORK_CLASSIFICATION','비가동 분류','SELECT DISTINCT','대분류 구분'],
    ['★★','MWIPLOTCQH','(품질관련 컬럼)','합격/불합격/보류','샘플 조회','품질 합격률'],
    ['★','MWIPORDSTS','ORD_CMF_1~30','커스텀 필드 (업체별 정의)','샘플 조회','추가 분석 차원'],
    ['★','MADMTBLDAT','(전체)','공통코드 테이블','SELECT *','코드값 해석 사전'],
    ['★','MWIPORDSTS','ORD_PRIORITY','우선순위','SELECT DISTINCT','긴급 지시 필터'],
]
for cd in codes:
    for c, v in enumerate(cd, 1):
        ws3.cell(row=r, column=c, value=v)
    style_row(ws3, r, 6)
    r += 1

r += 2
ws3.merge_cells(f'A{r}:F{r}')
ws3[f'A{r}'] = '※ 위 코드값들을 확인해야 정확한 실시간 대시보드 쿼리를 설계할 수 있습니다.'
ws3[f'A{r}'].font = Font(name='Arial', size=10, color='FF0000', bold=True)

for c in range(1, 7):
    ws3.column_dimensions[get_column_letter(c)].width = [10, 16, 24, 30, 30, 25][c-1]

# ═══════════════════════════════════════════════
# 시트4: 데이터 없는 테이블 (참고)
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('미사용 테이블(0건)')
ws4.sheet_properties.tabColor = 'D9D9D9'

ws4.merge_cells('A1:D1')
ws4['A1'] = '데이터 0건 테이블 (현재 미사용 - 참고용)'
ws4['A1'].font = title_font

r = 3
empty_headers = ['테이블명', '모듈', '추정 용도', '비고']
for c, h in enumerate(empty_headers, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 4, header_fill_master)
r += 1

empties = [
    ['CFMSORDSTS','설비보전','보전 작업지시 현황','설비보전 모듈 미사용'],
    ['CFMSORDHIS','설비보전','보전 작업지시 이력',''],
    ['CFMSORDSPT','설비보전','보전 Spare Part',''],
    ['CFMSORDTSK','설비보전','보전 작업 Task',''],
    ['CFMSORDECP','설비보전','보전 점검',''],
    ['CFMSORDETC','설비보전','보전 기타',''],
    ['CFMSPLNORD','설비보전','보전 계획별 지시',''],
    ['MINVSHPORD','MES코어','출하지시','출하지시 테이블 미사용'],
    ['MWIPCYTHIS','MES코어','사이클타임 이력','사이클타임 미등록'],
    ['MWIPCYTSTS','MES코어','사이클타임 현황',''],
    ['MFDCSUMLOT','MES코어','FDC LOT 집계','FDC 미사용'],
    ['MWIPLINWTM','MES코어','라인 가동시간','라인 가동시간 미등록'],
    ['MWIPLINOPH','MES코어','라인 운영이력',''],
]
for ed in empties:
    for c, v in enumerate(ed, 1):
        ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, 4)
    r += 1

for c in range(1, 5):
    ws4.column_dimensions[get_column_letter(c)].width = [18, 12, 25, 30][c-1]

output = 'C:/Users/user/Desktop/C&C/claude/MES_DB_테이블맵_분석.xlsx'
wb.save(output)
print(f'Saved: {output}')
