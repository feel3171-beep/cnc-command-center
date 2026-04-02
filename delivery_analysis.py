import pymssql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

conn = pymssql.connect(server='192.161.0.16', user='mestmp', password='cncmgr123!', database='MES', charset='utf8')

thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='C00000')
sub_fill = PatternFill('solid', fgColor='2F5496')
data_font = Font(name='맑은 고딕', size=10)
title_font = Font(name='맑은 고딕', bold=True, size=14, color='C00000')
section_font = Font(name='맑은 고딕', bold=True, size=12, color='2F5496')
alert_fill = PatternFill('solid', fgColor='FFC7CE')
alert_font = Font(name='맑은 고딕', bold=True, size=10, color='9C0006')
good_fill = PatternFill('solid', fgColor='C6EFCE')
warn_fill = PatternFill('solid', fgColor='FFEB9C')
risk_fill = PatternFill('solid', fgColor='FF6B6B')
risk_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')

def style_header(ws, row, max_col, fill=header_fill):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_all

def style_data(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = border_all
        cell.alignment = Alignment(vertical='center')

def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        mx = min_w
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) * 1.2 + 2, max_w))
        ws.column_dimensions[letter].width = mx

def fmt_date(d):
    s = str(d) if d else ''
    return f'{s[:4]}-{s[4:6]}-{s[6:]}' if len(s) == 8 else s

wb = Workbook()

# ============================================================
# Sheet 1: 납기 리스크 수주 (RISK SO)
# ============================================================
ws1 = wb.active
ws1.title = '1.납기리스크 수주'
ws1.sheet_properties.tabColor = 'C00000'

ws1.cell(row=1, column=1, value='납기 리스크 수주 현황').font = title_font
ws1.cell(row=2, column=1, value=f'기준일: {datetime.now().strftime("%Y-%m-%d %H:%M")} | 납기 내 미완료 수주를 리스크로 분류').font = Font(name='맑은 고딕', size=10, color='808080')

q1 = """
SELECT i.SO_NO, i.CUST_PO_NO, i.CUSTOMER_CODE,
    COUNT(DISTINCT i.ORDER_NO) AS total_orders,
    SUM(CASE WHEN o.ORD_STATUS = 'CLOSE' THEN 1 ELSE 0 END) AS closed_cnt,
    SUM(CASE WHEN o.ORD_STATUS IN ('WAIT','CONFIRM','PROCESS') THEN 1 ELSE 0 END) AS active_cnt,
    SUM(i.ORD_QTY) AS total_qty,
    SUM(ISNULL(o.ORD_OUT_QTY,0)) AS total_out,
    MIN(i.ORD_START_TIME) AS start_date,
    MAX(i.ORD_END_TIME) AS deadline,
    CASE WHEN SUM(i.ORD_QTY) > 0 THEN CAST(SUM(ISNULL(o.ORD_OUT_QTY,0))*100.0/SUM(i.ORD_QTY) AS DECIMAL(5,1)) ELSE 0 END AS achieve_pct,
    SUM(i.ORD_QTY) - SUM(ISNULL(o.ORD_OUT_QTY,0)) AS remaining_qty
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IS NOT NULL AND i.SO_NO <> ''
GROUP BY i.SO_NO, i.CUST_PO_NO, i.CUSTOMER_CODE
HAVING SUM(ISNULL(o.ORD_OUT_QTY,0)) < SUM(i.ORD_QTY) * 0.98
    AND MAX(i.ORD_END_TIME) <= '20260405'
ORDER BY MAX(i.ORD_END_TIME), SUM(i.ORD_QTY) - SUM(ISNULL(o.ORD_OUT_QTY,0)) DESC
"""
df1 = pd.read_sql(q1, conn)

r = 4
ws1.cell(row=r, column=1, value=f'[1] 미완료 수주 ({len(df1)}건) - 납기 4/5 이내, 달성률 98% 미만').font = section_font
r += 1
h = ['수주번호','고객PO/프로젝트','거래처코드','지시건수','마감건','진행건','지시수량','산출수량','잔여수량','달성률(%)','시작일','납기일','긴급도']
for i, hd in enumerate(h, 1):
    ws1.cell(row=r, column=i, value=hd)
style_header(ws1, r, len(h))

r += 1
today = datetime.now().strftime('%Y%m%d')
for _, row in df1.iterrows():
    ws1.cell(row=r, column=1, value=row['SO_NO'])
    ws1.cell(row=r, column=2, value=row['CUST_PO_NO'] or '')
    ws1.cell(row=r, column=3, value=row['CUSTOMER_CODE'] or '')
    ws1.cell(row=r, column=4, value=int(row['total_orders']))
    ws1.cell(row=r, column=5, value=int(row['closed_cnt']))
    ws1.cell(row=r, column=6, value=int(row['active_cnt']))
    ws1.cell(row=r, column=7, value=int(row['total_qty']))
    ws1.cell(row=r, column=8, value=int(row['total_out']))
    ws1.cell(row=r, column=9, value=int(row['remaining_qty']))
    ws1.cell(row=r, column=10, value=float(row['achieve_pct']))
    ws1.cell(row=r, column=11, value=fmt_date(row['start_date']))
    ws1.cell(row=r, column=12, value=fmt_date(row['deadline']))

    deadline = str(row['deadline'])
    ach = float(row['achieve_pct'])
    if deadline < today:
        urgency = 'OVERDUE'
        ws1.cell(row=r, column=13, value=urgency)
        ws1.cell(row=r, column=13).fill = risk_fill
        ws1.cell(row=r, column=13).font = risk_font
    elif deadline <= '20260331' and ach < 50:
        urgency = 'HIGH'
        ws1.cell(row=r, column=13, value=urgency)
        ws1.cell(row=r, column=13).fill = alert_fill
        ws1.cell(row=r, column=13).font = alert_font
    elif deadline <= '20260331':
        urgency = 'MEDIUM'
        ws1.cell(row=r, column=13, value=urgency)
        ws1.cell(row=r, column=13).fill = warn_fill
    else:
        urgency = 'LOW'
        ws1.cell(row=r, column=13, value=urgency)
        ws1.cell(row=r, column=13).fill = good_fill

    style_data(ws1, r, len(h))
    for c in [7,8,9]:
        ws1.cell(row=r, column=c).number_format = '#,##0'
    ws1.cell(row=r, column=10).number_format = '0.0'
    ws1.cell(row=r, column=13).alignment = Alignment(horizontal='center', vertical='center')
    r += 1

auto_width(ws1)

# ============================================================
# Sheet 2: 거래처별 수주 달성률
# ============================================================
ws2 = wb.create_sheet('2.거래처별 달성률')
ws2.sheet_properties.tabColor = '2F5496'

ws2.cell(row=1, column=1, value='거래처별 수주 달성률 (3월 확정분)').font = Font(name='맑은 고딕', bold=True, size=14, color='2F5496')

q2 = """
SELECT i.CUSTOMER_CODE, MIN(i.CUST_PO_NO) AS sample_po,
    COUNT(DISTINCT i.SO_NO) AS so_cnt, COUNT(DISTINCT i.ORDER_NO) AS ord_cnt,
    SUM(i.ORD_QTY) AS total_qty, SUM(ISNULL(o.ORD_OUT_QTY,0)) AS total_out,
    CASE WHEN SUM(i.ORD_QTY)>0 THEN CAST(SUM(ISNULL(o.ORD_OUT_QTY,0))*100.0/SUM(i.ORD_QTY) AS DECIMAL(5,1)) ELSE 0 END AS achieve_pct,
    SUM(i.ORD_QTY) - SUM(ISNULL(o.ORD_OUT_QTY,0)) AS remaining,
    MIN(i.ORD_END_TIME) AS earliest_deadline,
    MAX(i.ORD_END_TIME) AS latest_deadline
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IS NOT NULL AND i.SO_NO <> '' AND i.CONFIRM_DATE >= '20260301'
GROUP BY i.CUSTOMER_CODE
ORDER BY SUM(i.ORD_QTY) DESC
"""
df2 = pd.read_sql(q2, conn)

r = 3
h2 = ['거래처코드','참조PO','수주건수','지시건수','총지시수량','총산출수량','잔여수량','달성률(%)','최초납기','최종납기']
for i, hd in enumerate(h2, 1):
    ws2.cell(row=r, column=i, value=hd)
style_header(ws2, r, len(h2), fill=sub_fill)

r += 1
for _, row in df2.iterrows():
    ws2.cell(row=r, column=1, value=row['CUSTOMER_CODE'])
    ws2.cell(row=r, column=2, value=(row['sample_po'] or '')[:30])
    ws2.cell(row=r, column=3, value=int(row['so_cnt']))
    ws2.cell(row=r, column=4, value=int(row['ord_cnt']))
    ws2.cell(row=r, column=5, value=int(row['total_qty']))
    ws2.cell(row=r, column=6, value=int(row['total_out']))
    ws2.cell(row=r, column=7, value=int(row['remaining']))
    ws2.cell(row=r, column=8, value=float(row['achieve_pct']))
    ws2.cell(row=r, column=9, value=fmt_date(row['earliest_deadline']))
    ws2.cell(row=r, column=10, value=fmt_date(row['latest_deadline']))
    style_data(ws2, r, len(h2))
    for c in [5,6,7]:
        ws2.cell(row=r, column=c).number_format = '#,##0'
    ws2.cell(row=r, column=8).number_format = '0.0'
    ach = float(row['achieve_pct'])
    if ach >= 95:
        ws2.cell(row=r, column=8).fill = good_fill
    elif ach >= 70:
        ws2.cell(row=r, column=8).fill = warn_fill
    else:
        ws2.cell(row=r, column=8).fill = alert_fill
    r += 1

auto_width(ws2)

# ============================================================
# Sheet 3: 주요 수주 상세 추적
# ============================================================
ws3 = wb.create_sheet('3.주요수주 상세추적')
ws3.sheet_properties.tabColor = 'ED7D31'

ws3.cell(row=1, column=1, value='주요 수주 상세 추적 (수량 기준 TOP)').font = Font(name='맑은 고딕', bold=True, size=14, color='ED7D31')

q3 = """
SELECT i.SO_NO, i.CUST_PO_NO, i.CUSTOMER_CODE, i.FACTORY_CODE,
    i.ORDER_NO, i.MAT_CODE, i.LINE_CODE, i.ORD_QTY,
    i.ORD_START_TIME, i.ORD_END_TIME,
    o.ORD_STATUS, ISNULL(o.ORD_OUT_QTY,0) AS out_qty,
    CASE WHEN i.ORD_QTY > 0 THEN CAST(ISNULL(o.ORD_OUT_QTY,0)*100.0/i.ORD_QTY AS DECIMAL(5,1)) ELSE 0 END AS pct
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IN (
    SELECT TOP 10 SO_NO FROM IWIPORDSTS
    WHERE SO_NO IS NOT NULL AND SO_NO <> '' AND CONFIRM_DATE >= '20260301'
    GROUP BY SO_NO ORDER BY SUM(ORD_QTY) DESC
)
ORDER BY i.SO_NO, i.ORD_START_TIME, i.ORDER_NO
"""
df3 = pd.read_sql(q3, conn)

factory_names = {'1100':'퍼플','1200':'그린','1300':'3공장'}
r = 3
h3 = ['수주번호','고객PO','거래처','공장','지시번호','자재코드','라인','지시수량','산출수량','달성률(%)','MES상태','시작일','납기일']
for i, hd in enumerate(h3, 1):
    ws3.cell(row=r, column=i, value=hd)
style_header(ws3, r, len(h3), fill=PatternFill('solid', fgColor='ED7D31'))

r += 1
prev_so = None
for _, row in df3.iterrows():
    if row['SO_NO'] != prev_so and prev_so is not None:
        for c in range(1, len(h3)+1):
            ws3.cell(row=r, column=c).border = Border(top=Side(style='medium'))
    prev_so = row['SO_NO']

    ws3.cell(row=r, column=1, value=row['SO_NO'])
    ws3.cell(row=r, column=2, value=(row['CUST_PO_NO'] or '')[:35])
    ws3.cell(row=r, column=3, value=row['CUSTOMER_CODE'] or '')
    ws3.cell(row=r, column=4, value=factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE']))
    ws3.cell(row=r, column=5, value=row['ORDER_NO'])
    ws3.cell(row=r, column=6, value=row['MAT_CODE'])
    ws3.cell(row=r, column=7, value=row['LINE_CODE'])
    ws3.cell(row=r, column=8, value=int(row['ORD_QTY']))
    ws3.cell(row=r, column=9, value=int(row['out_qty']))
    ws3.cell(row=r, column=10, value=float(row['pct']))
    ws3.cell(row=r, column=11, value=row['ORD_STATUS'] or '')
    ws3.cell(row=r, column=12, value=fmt_date(row['ORD_START_TIME']))
    ws3.cell(row=r, column=13, value=fmt_date(row['ORD_END_TIME']))
    style_data(ws3, r, len(h3))
    ws3.cell(row=r, column=8).number_format = '#,##0'
    ws3.cell(row=r, column=9).number_format = '#,##0'
    ws3.cell(row=r, column=10).number_format = '0.0'

    status = row['ORD_STATUS'] or ''
    if status == 'CLOSE':
        ws3.cell(row=r, column=11).fill = good_fill
    elif status in ('WAIT','PROCESS'):
        ws3.cell(row=r, column=11).fill = warn_fill
    elif status == 'CONFIRM':
        ws3.cell(row=r, column=11).fill = PatternFill('solid', fgColor='D6E4F0')
    r += 1

auto_width(ws3)

# ============================================================
# Sheet 4: 공장별 납기 부하
# ============================================================
ws4 = wb.create_sheet('4.공장별 납기부하')
ws4.sheet_properties.tabColor = '7030A0'

ws4.cell(row=1, column=1, value='공장별 주간 납기 부하 분석').font = Font(name='맑은 고딕', bold=True, size=14, color='7030A0')

q4 = """
SELECT i.FACTORY_CODE,
    CASE
        WHEN i.ORD_END_TIME < '20260329' THEN 'A.이번주이전'
        WHEN i.ORD_END_TIME BETWEEN '20260329' AND '20260404' THEN 'B.이번주(3/29~4/4)'
        WHEN i.ORD_END_TIME BETWEEN '20260405' AND '20260411' THEN 'C.다음주(4/5~4/11)'
        WHEN i.ORD_END_TIME BETWEEN '20260412' AND '20260418' THEN 'D.2주후(4/12~4/18)'
        ELSE 'E.3주후이상'
    END AS week_group,
    COUNT(DISTINCT i.ORDER_NO) AS ord_cnt,
    SUM(i.ORD_QTY) AS total_qty,
    SUM(ISNULL(o.ORD_OUT_QTY,0)) AS total_out,
    SUM(i.ORD_QTY) - SUM(ISNULL(o.ORD_OUT_QTY,0)) AS remaining
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IS NOT NULL AND i.SO_NO <> ''
    AND o.ORD_STATUS NOT IN ('CLOSE')
    AND i.ORD_END_TIME >= '20260101'
GROUP BY i.FACTORY_CODE,
    CASE
        WHEN i.ORD_END_TIME < '20260329' THEN 'A.이번주이전'
        WHEN i.ORD_END_TIME BETWEEN '20260329' AND '20260404' THEN 'B.이번주(3/29~4/4)'
        WHEN i.ORD_END_TIME BETWEEN '20260405' AND '20260411' THEN 'C.다음주(4/5~4/11)'
        WHEN i.ORD_END_TIME BETWEEN '20260412' AND '20260418' THEN 'D.2주후(4/12~4/18)'
        ELSE 'E.3주후이상'
    END
ORDER BY i.FACTORY_CODE,
    CASE
        WHEN i.ORD_END_TIME < '20260329' THEN 'A.이번주이전'
        WHEN i.ORD_END_TIME BETWEEN '20260329' AND '20260404' THEN 'B.이번주(3/29~4/4)'
        WHEN i.ORD_END_TIME BETWEEN '20260405' AND '20260411' THEN 'C.다음주(4/5~4/11)'
        WHEN i.ORD_END_TIME BETWEEN '20260412' AND '20260418' THEN 'D.2주후(4/12~4/18)'
        ELSE 'E.3주후이상'
    END
"""
df4 = pd.read_sql(q4, conn)

r = 3
h4 = ['공장','기간','미완료 지시건수','총지시수량','총산출수량','잔여수량','잔여비율(%)']
for i, hd in enumerate(h4, 1):
    ws4.cell(row=r, column=i, value=hd)
style_header(ws4, r, len(h4), fill=PatternFill('solid', fgColor='7030A0'))

r += 1
for _, row in df4.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    rem_pct = float(row['remaining'])/float(row['total_qty'])*100 if float(row['total_qty']) > 0 else 0
    ws4.cell(row=r, column=1, value=fname)
    ws4.cell(row=r, column=2, value=row['week_group'])
    ws4.cell(row=r, column=3, value=int(row['ord_cnt']))
    ws4.cell(row=r, column=4, value=int(row['total_qty']))
    ws4.cell(row=r, column=5, value=int(row['total_out']))
    ws4.cell(row=r, column=6, value=int(row['remaining']))
    ws4.cell(row=r, column=7, value=round(rem_pct, 1))
    style_data(ws4, r, len(h4))
    for c in [4,5,6]:
        ws4.cell(row=r, column=c).number_format = '#,##0'
    ws4.cell(row=r, column=7).number_format = '0.0'
    wg = row['week_group']
    if 'A.' in wg:
        ws4.cell(row=r, column=2).fill = alert_fill
    elif 'B.' in wg:
        ws4.cell(row=r, column=2).fill = warn_fill
    r += 1

auto_width(ws4)

# ============================================================
# Sheet 5: SO-지시 매핑 전체 (raw data)
# ============================================================
ws5 = wb.create_sheet('5.SO전체데이터')
ws5.sheet_properties.tabColor = '808080'

ws5.cell(row=1, column=1, value='3월 수주 연계 전체 데이터 (Raw)').font = Font(name='맑은 고딕', bold=True, size=14, color='808080')

q5 = """
SELECT i.SO_NO, i.CUST_PO_NO, i.CUSTOMER_CODE, i.FACTORY_CODE,
    i.ORDER_NO, i.CONFIRM_DATE, i.MAT_CODE, i.LINE_CODE,
    i.ORD_QTY, i.ORD_START_TIME, i.ORD_END_TIME,
    o.ORD_STATUS, o.PLAN_DATE, ISNULL(o.ORD_OUT_QTY,0) AS out_qty,
    ISNULL(o.RCV_GOOD_QTY,0) AS good_qty, ISNULL(o.RCV_LOSS_QTY,0) AS loss_qty,
    CASE WHEN i.ORD_QTY > 0 THEN CAST(ISNULL(o.ORD_OUT_QTY,0)*100.0/i.ORD_QTY AS DECIMAL(5,1)) ELSE 0 END AS pct
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IS NOT NULL AND i.SO_NO <> '' AND i.CONFIRM_DATE >= '20260301'
ORDER BY i.SO_NO, i.ORDER_NO
"""
df5 = pd.read_sql(q5, conn)

r = 3
h5 = ['수주번호','고객PO','거래처','공장','지시번호','확정일','자재코드','라인',
      '지시수량','시작일','납기일','MES상태','계획일','산출수량','양품수량','불량수량','달성률(%)']
for i, hd in enumerate(h5, 1):
    ws5.cell(row=r, column=i, value=hd)
style_header(ws5, r, len(h5), fill=PatternFill('solid', fgColor='808080'))

r += 1
for _, row in df5.iterrows():
    ws5.cell(row=r, column=1, value=row['SO_NO'])
    ws5.cell(row=r, column=2, value=(row['CUST_PO_NO'] or '')[:30])
    ws5.cell(row=r, column=3, value=row['CUSTOMER_CODE'] or '')
    ws5.cell(row=r, column=4, value=factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE']))
    ws5.cell(row=r, column=5, value=row['ORDER_NO'])
    ws5.cell(row=r, column=6, value=fmt_date(row['CONFIRM_DATE']))
    ws5.cell(row=r, column=7, value=row['MAT_CODE'])
    ws5.cell(row=r, column=8, value=row['LINE_CODE'])
    ws5.cell(row=r, column=9, value=int(row['ORD_QTY']))
    ws5.cell(row=r, column=10, value=fmt_date(row['ORD_START_TIME']))
    ws5.cell(row=r, column=11, value=fmt_date(row['ORD_END_TIME']))
    ws5.cell(row=r, column=12, value=row['ORD_STATUS'] or '')
    ws5.cell(row=r, column=13, value=fmt_date(row['PLAN_DATE']))
    ws5.cell(row=r, column=14, value=int(row['out_qty']))
    ws5.cell(row=r, column=15, value=int(row['good_qty']))
    ws5.cell(row=r, column=16, value=int(row['loss_qty']))
    ws5.cell(row=r, column=17, value=float(row['pct']))
    style_data(ws5, r, len(h5))
    for c in [9,14,15,16]:
        ws5.cell(row=r, column=c).number_format = '#,##0'
    ws5.cell(row=r, column=17).number_format = '0.0'
    r += 1

auto_width(ws5)

out_path = 'C:/Users/user/Desktop/C&C/claude/MES_납기분석_대시보드.xlsx'
wb.save(out_path)
conn.close()
print(f'Saved: {out_path}')
print(f'Sheets: {wb.sheetnames}')
print(f'Sheet1 리스크수주: {len(df1)}건')
print(f'Sheet2 거래처: {len(df2)}건')
print(f'Sheet3 주요수주상세: {len(df3)}건')
print(f'Sheet5 전체raw: {len(df5)}건')
