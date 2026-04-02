from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
data_font = Font(name='Arial', size=10)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.alignment = center
        cell.border = border

# === 시트1: 생산계획 ===
ws1 = wb.active
ws1.title = "생산계획"
ws1.append(["공장", "주간계획수량", "비고"])
style_header(ws1, 1, 3)

plan_data = [
    ["퍼플", 12000, "정상 가동"],
    ["그린", 8500, "야간 포함"],
    ["3공장", 6200, ""],
    ["외주", 3800, "A업체"],
]
for row_data in plan_data:
    ws1.append(row_data)
for r in range(2, 6):
    style_data(ws1, r, 3)

ws1.append(["합계", None, ""])
ws1.cell(row=6, column=2, value="=SUM(B2:B5)")
style_data(ws1, 6, 3)
ws1.cell(row=6, column=1).font = Font(name='Arial', bold=True, size=10)
ws1.cell(row=6, column=2).font = Font(name='Arial', bold=True, size=10)

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 15

# === 시트2: 일별생산량 ===
ws2 = wb.create_sheet("일별생산량")
ws2.append(["날짜", "퍼플", "그린", "3공장", "외주", "합계"])
style_header(ws2, 1, 6)

daily_data = [
    ["2026-03-09", 2450, 1720, 1280, 780, None],
    ["2026-03-10", 2380, 1690, 1250, 760, None],
    ["2026-03-11", 2520, 1750, 1300, 790, None],
    ["2026-03-12", 2400, 1700, 1260, 770, None],
    ["2026-03-13", 2480, 1730, 1290, 785, None],
]
for i, row_data in enumerate(daily_data):
    ws2.append(row_data)
    r = i + 2
    ws2.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
    style_data(ws2, r, 6)

ws2.append(["누적합계", None, None, None, None, None])
r = 7
for c in range(2, 7):
    col = get_column_letter(c)
    ws2.cell(row=r, column=c, value=f"=SUM({col}2:{col}6)")
style_data(ws2, r, 6)
ws2.cell(row=r, column=1).font = Font(name='Arial', bold=True, size=10)

ws2.column_dimensions['A'].width = 14
for c in range(2, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 12

# === 시트3: 인원근무 ===
ws3 = wb.create_sheet("인원근무")
ws3.append(["날짜", "투입인원(명)", "평균근무시간(h)"])
style_header(ws3, 1, 3)

hr_data = [
    ["2026-03-09", 128, 9.2],
    ["2026-03-10", 125, 9.0],
    ["2026-03-11", 130, 9.5],
    ["2026-03-12", 127, 9.1],
    ["2026-03-13", 129, 9.3],
]
for row_data in hr_data:
    ws3.append(row_data)
for r in range(2, 7):
    style_data(ws3, r, 3)

ws3.append(["평균", None, None])
ws3.cell(row=7, column=2, value="=AVERAGE(B2:B6)")
ws3.cell(row=7, column=3, value="=AVERAGE(C2:C6)")
style_data(ws3, 7, 3)
ws3.cell(row=7, column=1).font = Font(name='Arial', bold=True, size=10)

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 18

output = "C:/Users/user/Desktop/C&C/claude/slack_bot/sap_data/생산현황_템플릿.xlsx"
wb.save(output)
print(f"템플릿 생성 완료: {output}")
