import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb['매출채권_RAW']

# Check which columns have actual data (sample rows 2-100)
cols_with_data = []
for col in range(1, 839):
    has_data = False
    for r in range(2, min(102, ws.max_row + 1)):
        c = ws.cell(r, col)
        if c.value is not None and str(c.value).strip() not in ('', '0'):
            has_data = True
            break
    if has_data:
        cl = openpyxl.utils.get_column_letter(col)
        header = ws.cell(1, col).value or ''
        cols_with_data.append((col, cl, str(header)[:40]))

print(f"838열 중 실제 데이터가 있는 열: {len(cols_with_data)}개")
for col_num, cl, header in cols_with_data:
    print(f"  {cl} ({col_num}): {header}")

# Also check which columns are referenced by ⑪ 매출채권 (OUTPUT)
wb2 = openpyxl.load_workbook(file)
ws_out = wb2['⑪ 매출채권 (OUTPUT)']
referenced_cols = set()
for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row, max_col=ws_out.max_column):
    for c in row:
        if c.data_type == 'f' and c.value and '매출채권_RAW' in str(c.value):
            formula = str(c.value)
            # Extract column references like $F:$F, $A:$A, $ML:$ML
            import re
            refs = re.findall(r'\$([A-Z]+):', formula)
            referenced_cols.update(refs)

print(f"\n⑪ 매출채권 (OUTPUT)에서 참조하는 매출채권_RAW 열: {sorted(referenced_cols)}")
