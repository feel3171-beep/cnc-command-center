import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
wb = openpyxl.load_workbook(file)

print('=== Sheets ===')
for i, s in enumerate(wb.sheetnames):
    print(f'  [{i}] {s}')

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
        vals = []
        for c in row:
            if c.value is not None:
                prefix = '[F]' if c.data_type == 'f' else ''
                v = str(c.value)
                if len(v) > 60:
                    v = v[:60] + '...'
                vals.append(f'{c.coordinate}={prefix}{v}')
        if vals:
            print(' | '.join(vals))
