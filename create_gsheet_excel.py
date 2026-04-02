import pandas as pd
import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
output = r'C:\Users\user\Downloads\FY2602_결산_GSheet용.xlsx'

wb_orig = openpyxl.load_workbook(file, data_only=True)
writer = pd.ExcelWriter(output, engine='openpyxl')

for sname in wb_orig.sheetnames:
    print(f"Processing: {sname}...")
    ws = wb_orig[sname]

    if sname == '생산량_RAW':
        max_row = 50000
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column, values_only=True):
            data.append(list(row))
        if data:
            df = pd.DataFrame(data[1:], columns=data[0])
            df = df.dropna(subset=[df.columns[6], df.columns[7], df.columns[14]], how='all')
            df.to_excel(writer, sheet_name=sname, index=False)
            print(f"  -> {len(df)} rows x {len(df.columns)} cols")

    elif sname == '매출채권_RAW':
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            data.append(list(row))
        if data:
            df = pd.DataFrame(data[1:])  # Use numeric index for columns
            # Find non-empty columns by index
            non_empty_idx = []
            for i in range(len(df.columns)):
                col = df.iloc[:, i]
                if col.notna().any():
                    non_zero = col.dropna()
                    if len(non_zero) > 0 and not (non_zero.astype(str).str.strip() == '').all():
                        non_empty_idx.append(i)
            df = df.iloc[:, non_empty_idx]
            # Set header names from original
            headers = [data[0][i] for i in non_empty_idx]
            # Make headers unique
            seen = {}
            unique_headers = []
            for h in headers:
                h_str = str(h) if h else ''
                if h_str in seen:
                    seen[h_str] += 1
                    unique_headers.append(f"{h_str}_{seen[h_str]}")
                else:
                    seen[h_str] = 0
                    unique_headers.append(h_str)
            df.columns = unique_headers
            df = df.dropna(how='all')
            df.to_excel(writer, sheet_name=sname, index=False)
            print(f"  -> {len(df)} rows x {len(df.columns)} cols")

    else:
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            data.append(list(row))
        if data and len(data) > 1:
            df = pd.DataFrame(data[1:], columns=data[0])
            df = df.dropna(how='all')
            df.to_excel(writer, sheet_name=sname, index=False)
            print(f"  -> {len(df)} rows x {len(df.columns)} cols")

writer.close()
print(f"\nDone! Output: {output}")

import os
size_mb = os.path.getsize(output) / (1024 * 1024)
print(f"File size: {size_mb:.1f} MB")
