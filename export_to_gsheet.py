import openpyxl
import pandas as pd
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'

# Get sheet info for planning
wb = openpyxl.load_workbook(file, data_only=True)

sheets_info = []
for sname in wb.sheetnames:
    ws = wb[sname]
    actual_max = ws.max_row

    # For 생산량_RAW, find actual data end
    if sname == '생산량_RAW':
        actual_max = 50000  # Confirmed data up to ~49999

    sheets_info.append({
        'name': sname,
        'rows': actual_max,
        'cols': ws.max_column,
        'cells': actual_max * ws.max_column
    })
    print(f"{sname}: {actual_max} rows x {ws.max_column} cols = {actual_max * ws.max_column:,} cells")

total_cells = sum(s['cells'] for s in sheets_info)
print(f"\nTotal cells: {total_cells:,}")
print(f"Google Sheets limit: 10,000,000 cells")
print(f"{'OK' if total_cells < 10000000 else 'OVER LIMIT'}")

# Check which sheets are over the limit individually
for s in sheets_info:
    if s['cells'] > 5000000:
        print(f"  WARNING: {s['name']} alone has {s['cells']:,} cells")
