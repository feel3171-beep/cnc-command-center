import pymssql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

conn = pymssql.connect(server='192.161.0.16', user='mestmp', password='cncmgr123!', database='MES', charset='utf8')

thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
sub_header_fill = PatternFill('solid', fgColor='4472C4')
data_font = Font(name='맑은 고딕', size=10)
title_font = Font(name='맑은 고딕', bold=True, size=14, color='2F5496')
section_font = Font(name='맑은 고딕', bold=True, size=12, color='2F5496')
alert_fill = PatternFill('solid', fgColor='FFC7CE')
good_fill = PatternFill('solid', fgColor='C6EFCE')
warn_fill = PatternFill('solid', fgColor='FFEB9C')
light_blue = PatternFill('solid', fgColor='D6E4F0')

def style_header(ws, row, max_col, fill=header_fill):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_all

def style_data(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = border_all
        cell.alignment = Alignment(vertical='center')

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx = min_w
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) * 1.2 + 2, max_w))
        ws.column_dimensions[letter].width = mx

wb = Workbook()

# ============================================================
# Sheet 1: 작업지시 현황 (Order Status Dashboard)
# ============================================================
ws1 = wb.active
ws1.title = '1.작업지시 현황'
ws1.sheet_properties.tabColor = '2F5496'

ws1.cell(row=1, column=1, value='작업지시 실시간 현황 대시보드').font = title_font
ws1.cell(row=2, column=1, value=f'조회시점: {datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(name='맑은 고딕', size=10, color='808080')

# 1-1. 공장별 상태 요약
ws1.cell(row=4, column=1, value='[1] 공장별 작업지시 상태 요약').font = section_font
headers = ['공장','WAIT(대기)','WAIT수량','CONFIRM(확정)','CONFIRM수량','PROCESS(진행)','PROCESS수량','CLOSE(마감)','CLOSE수량']
r = 5
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, len(headers))

q = """
SELECT FACTORY_CODE,
    SUM(CASE WHEN ORD_STATUS='WAIT' THEN 1 ELSE 0 END) AS wait_cnt,
    SUM(CASE WHEN ORD_STATUS='WAIT' THEN ORD_QTY ELSE 0 END) AS wait_qty,
    SUM(CASE WHEN ORD_STATUS='CONFIRM' THEN 1 ELSE 0 END) AS cfm_cnt,
    SUM(CASE WHEN ORD_STATUS='CONFIRM' THEN ORD_QTY ELSE 0 END) AS cfm_qty,
    SUM(CASE WHEN ORD_STATUS='PROCESS' THEN 1 ELSE 0 END) AS proc_cnt,
    SUM(CASE WHEN ORD_STATUS='PROCESS' THEN ORD_QTY ELSE 0 END) AS proc_qty,
    SUM(CASE WHEN ORD_STATUS='CLOSE' THEN 1 ELSE 0 END) AS close_cnt,
    SUM(CASE WHEN ORD_STATUS='CLOSE' THEN ORD_QTY ELSE 0 END) AS close_qty
FROM MWIPORDSTS
WHERE ORD_STATUS NOT IN ('DELETE')
GROUP BY FACTORY_CODE
ORDER BY FACTORY_CODE
"""
df = pd.read_sql(q, conn)
factory_names = {'1100':'퍼플카운티','1200':'그린카운티','1300':'3공장'}
r = 6
for _, row in df.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    ws1.cell(row=r, column=1, value=fname)
    ws1.cell(row=r, column=2, value=int(row['wait_cnt']))
    ws1.cell(row=r, column=3, value=int(row['wait_qty']))
    ws1.cell(row=r, column=4, value=int(row['cfm_cnt']))
    ws1.cell(row=r, column=5, value=int(row['cfm_qty']))
    ws1.cell(row=r, column=6, value=int(row['proc_cnt']))
    ws1.cell(row=r, column=7, value=int(row['proc_qty']))
    ws1.cell(row=r, column=8, value=int(row['close_cnt']))
    ws1.cell(row=r, column=9, value=int(row['close_qty']))
    style_data(ws1, r, len(headers))
    for c in [3,5,7,9]:
        ws1.cell(row=r, column=c).number_format = '#,##0'
    r += 1

# 1-2. 진행중 + 대기 작업 상세 (최근 계획일 기준)
r += 2
ws1.cell(row=r, column=1, value='[2] 미완료 작업지시 상세 (WAIT/CONFIRM/PROCESS - 최근 계획일 순)').font = section_font
r += 1
headers2 = ['공장','지시번호','계획일','상태','자재코드','라인','지시수량','투입수량','산출수량','진행률(%)']
for i, h in enumerate(headers2, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, len(headers2))

q2 = """
SELECT TOP 50 FACTORY_CODE, ORDER_NO, PLAN_DATE, ORD_STATUS, MAT_CODE, LINE_CODE,
    ORD_QTY, ORD_IN_QTY, ORD_OUT_QTY,
    CASE WHEN ORD_QTY > 0 THEN CAST(ORD_OUT_QTY * 100.0 / ORD_QTY AS DECIMAL(5,1)) ELSE 0 END AS progress
FROM MWIPORDSTS
WHERE ORD_STATUS IN ('WAIT','CONFIRM','PROCESS')
ORDER BY PLAN_DATE DESC, FACTORY_CODE
"""
df2 = pd.read_sql(q2, conn)
r += 1
for _, row in df2.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    ws1.cell(row=r, column=1, value=fname)
    ws1.cell(row=r, column=2, value=row['ORDER_NO'])
    plan_d = str(row['PLAN_DATE'])
    ws1.cell(row=r, column=3, value=f'{plan_d[:4]}-{plan_d[4:6]}-{plan_d[6:]}' if len(plan_d)==8 else plan_d)
    ws1.cell(row=r, column=4, value=row['ORD_STATUS'])
    ws1.cell(row=r, column=5, value=row['MAT_CODE'])
    ws1.cell(row=r, column=6, value=row['LINE_CODE'])
    ws1.cell(row=r, column=7, value=int(row['ORD_QTY']))
    ws1.cell(row=r, column=8, value=int(row['ORD_IN_QTY']))
    ws1.cell(row=r, column=9, value=int(row['ORD_OUT_QTY']))
    ws1.cell(row=r, column=10, value=float(row['progress']))
    style_data(ws1, r, len(headers2))
    ws1.cell(row=r, column=7).number_format = '#,##0'
    ws1.cell(row=r, column=8).number_format = '#,##0'
    ws1.cell(row=r, column=9).number_format = '#,##0'
    ws1.cell(row=r, column=10).number_format = '0.0'
    # 색상: PROCESS=초록, WAIT=노랑, CONFIRM=파랑
    status = row['ORD_STATUS']
    if status == 'PROCESS':
        ws1.cell(row=r, column=4).fill = good_fill
    elif status == 'WAIT':
        ws1.cell(row=r, column=4).fill = warn_fill
    elif status == 'CONFIRM':
        ws1.cell(row=r, column=4).fill = light_blue
    # 진행률 > 90% 강조
    prog = float(row['progress'])
    if prog >= 90:
        ws1.cell(row=r, column=10).fill = good_fill
    elif prog > 0:
        ws1.cell(row=r, column=10).fill = warn_fill
    r += 1

auto_width(ws1)

# ============================================================
# Sheet 2: 라인별 생산 실적 (Production by Line)
# ============================================================
ws2 = wb.create_sheet('2.라인별 생산실적')
ws2.sheet_properties.tabColor = '00B050'

ws2.cell(row=1, column=1, value='라인별 일자별 생산 실적 (3월)').font = title_font
ws2.cell(row=2, column=1, value=f'데이터 소스: MWIPLOTHIS (TRAN_CODE=CV, 생산완료)').font = Font(name='맑은 고딕', size=10, color='808080')

# 공장별 라인별 일자별 생산수량
q3 = """
SELECT l.FACTORY_CODE, l.LINE_CODE, l.LINE_DESC, l.LINE_TYPE,
    SUBSTRING(CONVERT(VARCHAR, h.TRAN_TIME, 112), 7, 2) AS day_str,
    SUM(h.QTY) AS daily_qty
FROM MWIPLOTHIS h
JOIN MWIPLINDEF l ON h.FACTORY_CODE = l.FACTORY_CODE AND h.LINE_CODE = l.LINE_CODE
WHERE h.TRAN_CODE = 'CV'
    AND h.TRAN_TIME >= '2026-03-01' AND h.TRAN_TIME < '2026-03-30'
GROUP BY l.FACTORY_CODE, l.LINE_CODE, l.LINE_DESC, l.LINE_TYPE,
    SUBSTRING(CONVERT(VARCHAR, h.TRAN_TIME, 112), 7, 2)
"""
df3 = pd.read_sql(q3, conn)

# 공장별 라인별 총합
q3b = """
SELECT l.FACTORY_CODE, l.LINE_CODE, l.LINE_DESC, l.LINE_TYPE,
    SUM(h.QTY) AS total_qty, COUNT(DISTINCT CONVERT(VARCHAR, h.TRAN_TIME, 112)) AS work_days
FROM MWIPLOTHIS h
JOIN MWIPLINDEF l ON h.FACTORY_CODE = l.FACTORY_CODE AND h.LINE_CODE = l.LINE_CODE
WHERE h.TRAN_CODE = 'CV'
    AND h.TRAN_TIME >= '2026-03-01' AND h.TRAN_TIME < '2026-03-30'
GROUP BY l.FACTORY_CODE, l.LINE_CODE, l.LINE_DESC, l.LINE_TYPE
ORDER BY l.FACTORY_CODE, SUM(h.QTY) DESC
"""
df3b = pd.read_sql(q3b, conn)

r = 4
ws2.cell(row=r, column=1, value='[1] 공장별 라인별 3월 생산 실적 요약').font = section_font
r += 1
headers3 = ['공장','라인코드','라인명','라인유형','총생산수량','가동일수','일평균']
for i, h in enumerate(headers3, 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, len(headers3))

r += 1
for _, row in df3b.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    ws2.cell(row=r, column=1, value=fname)
    ws2.cell(row=r, column=2, value=row['LINE_CODE'])
    ws2.cell(row=r, column=3, value=row['LINE_DESC'])
    ws2.cell(row=r, column=4, value=row['LINE_TYPE'])
    ws2.cell(row=r, column=5, value=int(row['total_qty']))
    ws2.cell(row=r, column=6, value=int(row['work_days']))
    avg = int(row['total_qty']) / int(row['work_days']) if int(row['work_days']) > 0 else 0
    ws2.cell(row=r, column=7, value=round(avg))
    style_data(ws2, r, len(headers3))
    ws2.cell(row=r, column=5).number_format = '#,##0'
    ws2.cell(row=r, column=7).number_format = '#,##0'
    r += 1

auto_width(ws2)

# ============================================================
# Sheet 3: 비가동 분석 (Non-work Analysis)
# ============================================================
ws3 = wb.create_sheet('3.비가동 분석')
ws3.sheet_properties.tabColor = 'FF0000'

ws3.cell(row=1, column=1, value='라인별 비가동 시간 분석 (2026년)').font = title_font

nwk_code_map = {
    'E101':'작업준비','E102':'품목교체(호수)','E103':'품목교체(제품)',
    'E104':'라벨러(라벨부착기)','E105':'포장사양컨펌대기',
    'E201':'충전부 고장','E202':'캡핑부 고장','E203':'컨베이어 고장',
    'E204':'PM 셋팅 지연','E205':'기타 고장','E206':'순간정비',
    'E207':'착인기','E208':'설비 기타 고장',
    'E301':'자재 불량 확인 대기','E302':'자재 불량 환입',
    'E303':'자재불출 지연','E304':'품질이슈',
    'E401':'QC 확인대기','E402':'표준서(관리품) 대기',
    'E403':'내용물 불량 확인 대기','E404':'미품이슈',
    'E501':'벌크 보충','E502':'벌크 불량 환입',
    'E503':'탈포,재믹싱','E504':'멜팅,경도,컨펌대기',
    'E601':'청소정리','E602':'교육/회의','E603':'생산조건 조정',
    'E405':'(미등록)E405','E406':'(미등록)E406',
    'E505':'(미등록)E505','E506':'(미등록)E506',
    'B101':'(미등록)B101','C102':'(미등록)C102','D201':'(미등록)D201',
    'D501':'(미등록)D501'
}

nwk_class_map = {
    'E1':'작업교체','E2':'설비고장','E3':'자재이슈','E4':'품질대기','E5':'벌크이슈','E6':'기타'
}

q4 = """
SELECT FACTORY_CODE, NONWORK_CLASSIFICATION, NONWORK_CODE,
    COUNT(*) AS cnt, SUM(NONWORK_SECOND) AS total_sec, SUM(PURE_NONWORK_SECOND) AS pure_sec
FROM MWIPNWKSTS
WHERE NONWORK_DATE >= '20260301'
GROUP BY FACTORY_CODE, NONWORK_CLASSIFICATION, NONWORK_CODE
ORDER BY FACTORY_CODE, SUM(NONWORK_SECOND) DESC
"""
df4 = pd.read_sql(q4, conn)

r = 3
ws3.cell(row=r, column=1, value='[1] 공장별 비가동 코드별 시간 (3월~현재)').font = section_font
r += 1
headers4 = ['공장','분류','비가동코드','코드명','건수','총 비가동(시간)','순수 비가동(시간)','건당 평균(분)']
for i, h in enumerate(headers4, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, len(headers4))

r += 1
for _, row in df4.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    cls = nwk_class_map.get(row['NONWORK_CLASSIFICATION'], row['NONWORK_CLASSIFICATION'])
    code_desc = nwk_code_map.get(row['NONWORK_CODE'], row['NONWORK_CODE'])
    total_hrs = int(row['total_sec']) / 3600
    pure_hrs = int(row['pure_sec']) / 3600
    avg_min = int(row['total_sec']) / int(row['cnt']) / 60 if int(row['cnt']) > 0 else 0

    ws3.cell(row=r, column=1, value=fname)
    ws3.cell(row=r, column=2, value=cls)
    ws3.cell(row=r, column=3, value=row['NONWORK_CODE'])
    ws3.cell(row=r, column=4, value=code_desc)
    ws3.cell(row=r, column=5, value=int(row['cnt']))
    ws3.cell(row=r, column=6, value=round(total_hrs, 1))
    ws3.cell(row=r, column=7, value=round(pure_hrs, 1))
    ws3.cell(row=r, column=8, value=round(avg_min, 1))
    style_data(ws3, r, len(headers4))
    # 설비고장 빨간 강조
    if row['NONWORK_CLASSIFICATION'] == 'E2':
        ws3.cell(row=r, column=2).fill = alert_fill
    r += 1

# 라인별 비가동 TOP
r += 2
ws3.cell(row=r, column=1, value='[2] 라인별 비가동 시간 TOP 20 (3월~현재)').font = section_font
r += 1

q5 = """
SELECT TOP 20 n.FACTORY_CODE, n.LINE_CODE, l.LINE_DESC,
    COUNT(*) AS cnt, SUM(n.NONWORK_SECOND) AS total_sec,
    SUM(n.PURE_NONWORK_SECOND) AS pure_sec
FROM MWIPNWKSTS n
LEFT JOIN MWIPLINDEF l ON n.FACTORY_CODE = l.FACTORY_CODE AND n.LINE_CODE = l.LINE_CODE
WHERE n.NONWORK_DATE >= '20260301'
GROUP BY n.FACTORY_CODE, n.LINE_CODE, l.LINE_DESC
ORDER BY SUM(n.NONWORK_SECOND) DESC
"""
df5 = pd.read_sql(q5, conn)

headers5 = ['공장','라인코드','라인명','비가동 건수','총 비가동(시간)','순수 비가동(시간)']
for i, h in enumerate(headers5, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, len(headers5))
r += 1

for _, row in df5.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    ws3.cell(row=r, column=1, value=fname)
    ws3.cell(row=r, column=2, value=row['LINE_CODE'])
    ws3.cell(row=r, column=3, value=row['LINE_DESC'] or '')
    ws3.cell(row=r, column=4, value=int(row['cnt']))
    ws3.cell(row=r, column=5, value=round(int(row['total_sec'])/3600, 1))
    ws3.cell(row=r, column=6, value=round(int(row['pure_sec'])/3600, 1))
    style_data(ws3, r, len(headers5))
    r += 1

auto_width(ws3)

# ============================================================
# Sheet 4: 일별 생산 추이 (Daily Production Trend)
# ============================================================
ws4 = wb.create_sheet('4.일별 생산추이')
ws4.sheet_properties.tabColor = 'FFC000'

ws4.cell(row=1, column=1, value='공장별 일별 생산수량 추이 (3월)').font = title_font

q6 = """
SELECT CONVERT(VARCHAR, TRAN_TIME, 23) AS tran_day, FACTORY_CODE, SUM(QTY) AS daily_qty
FROM MWIPLOTHIS
WHERE TRAN_CODE = 'CV' AND TRAN_TIME >= '2026-03-01' AND TRAN_TIME < '2026-03-30'
GROUP BY CONVERT(VARCHAR, TRAN_TIME, 23), FACTORY_CODE
ORDER BY CONVERT(VARCHAR, TRAN_TIME, 23), FACTORY_CODE
"""
df6 = pd.read_sql(q6, conn)

pivot = df6.pivot_table(values='daily_qty', index='tran_day', columns='FACTORY_CODE', aggfunc='sum', fill_value=0)
pivot['합계'] = pivot.sum(axis=1)

r = 3
headers6 = ['날짜','퍼플카운티(1100)','그린카운티(1200)','3공장(1300)','합계']
for i, h in enumerate(headers6, 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, len(headers6))

r += 1
for day, prow in pivot.iterrows():
    ws4.cell(row=r, column=1, value=day)
    ws4.cell(row=r, column=2, value=int(prow.get('1100', 0)))
    ws4.cell(row=r, column=3, value=int(prow.get('1200', 0)))
    ws4.cell(row=r, column=4, value=int(prow.get('1300', 0)))
    ws4.cell(row=r, column=5, value=int(prow['합계']))
    style_data(ws4, r, len(headers6))
    for c in range(2, 6):
        ws4.cell(row=r, column=c).number_format = '#,##0'
    r += 1

# 합계행
ws4.cell(row=r, column=1, value='합계')
ws4.cell(row=r, column=1).font = Font(name='맑은 고딕', bold=True, size=10)
for c_idx, code in enumerate(['1100','1200','1300'], 2):
    ws4.cell(row=r, column=c_idx, value=int(pivot[code].sum()) if code in pivot.columns else 0)
    ws4.cell(row=r, column=c_idx).number_format = '#,##0'
    ws4.cell(row=r, column=c_idx).font = Font(name='맑은 고딕', bold=True, size=10)
ws4.cell(row=r, column=5, value=int(pivot['합계'].sum()))
ws4.cell(row=r, column=5).number_format = '#,##0'
ws4.cell(row=r, column=5).font = Font(name='맑은 고딕', bold=True, size=10)
style_data(ws4, r, len(headers6))

auto_width(ws4)

# ============================================================
# Sheet 5: 지시 대비 실적 (Order vs Actual)
# ============================================================
ws5 = wb.create_sheet('5.지시대비실적')
ws5.sheet_properties.tabColor = '7030A0'

ws5.cell(row=1, column=1, value='작업지시 대비 실적 분석 (CLOSE 지시 - 3월)').font = title_font

q7 = """
SELECT o.FACTORY_CODE, o.ORDER_NO, o.PLAN_DATE, o.MAT_CODE, o.LINE_CODE,
    o.ORD_QTY, o.ORD_OUT_QTY, o.RCV_GOOD_QTY, o.RCV_LOSS_QTY,
    CASE WHEN o.ORD_QTY > 0 THEN CAST(o.ORD_OUT_QTY * 100.0 / o.ORD_QTY AS DECIMAL(5,1)) ELSE 0 END AS achieve_rate,
    CASE WHEN o.ORD_OUT_QTY > 0 THEN CAST(o.RCV_LOSS_QTY * 100.0 / o.ORD_OUT_QTY AS DECIMAL(5,2)) ELSE 0 END AS loss_rate
FROM MWIPORDSTS o
WHERE o.ORD_STATUS = 'CLOSE' AND o.PLAN_DATE >= '20260301' AND o.PLAN_DATE < '20260330'
ORDER BY o.PLAN_DATE DESC, o.FACTORY_CODE
"""
df7 = pd.read_sql(q7, conn)

r = 3
headers7 = ['공장','지시번호','계획일','자재코드','라인','지시수량','산출수량','양품수량','불량수량','달성률(%)','불량률(%)']
for i, h in enumerate(headers7, 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, len(headers7))

r += 1
for _, row in df7.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    plan_d = str(row['PLAN_DATE'])
    ws5.cell(row=r, column=1, value=fname)
    ws5.cell(row=r, column=2, value=row['ORDER_NO'])
    ws5.cell(row=r, column=3, value=f'{plan_d[:4]}-{plan_d[4:6]}-{plan_d[6:]}' if len(plan_d)==8 else plan_d)
    ws5.cell(row=r, column=4, value=row['MAT_CODE'])
    ws5.cell(row=r, column=5, value=row['LINE_CODE'])
    ws5.cell(row=r, column=6, value=int(row['ORD_QTY']))
    ws5.cell(row=r, column=7, value=int(row['ORD_OUT_QTY']))
    ws5.cell(row=r, column=8, value=int(row['RCV_GOOD_QTY']))
    ws5.cell(row=r, column=9, value=int(row['RCV_LOSS_QTY']))
    ws5.cell(row=r, column=10, value=float(row['achieve_rate']))
    ws5.cell(row=r, column=11, value=float(row['loss_rate']))
    style_data(ws5, r, len(headers7))
    for c in [6,7,8,9]:
        ws5.cell(row=r, column=c).number_format = '#,##0'
    ws5.cell(row=r, column=10).number_format = '0.0'
    ws5.cell(row=r, column=11).number_format = '0.00'
    # 달성률 < 95% 경고
    if float(row['achieve_rate']) < 95 and float(row['achieve_rate']) > 0:
        ws5.cell(row=r, column=10).fill = alert_fill
    elif float(row['achieve_rate']) >= 100:
        ws5.cell(row=r, column=10).fill = good_fill
    # 불량률 > 3% 경고
    if float(row['loss_rate']) > 3:
        ws5.cell(row=r, column=11).fill = alert_fill
    r += 1

auto_width(ws5)

# ============================================================
# Sheet 6: ERP 인터페이스/납품 연계
# ============================================================
ws6 = wb.create_sheet('6.ERP연계(SO)')
ws6.sheet_properties.tabColor = 'ED7D31'

ws6.cell(row=1, column=1, value='ERP 작업지시 인터페이스 (IWIPORDSTS - Sales Order 연계)').font = title_font
ws6.cell(row=2, column=1, value='SO_NO: ERP 수주번호, CUSTOMER_CODE: 거래처코드, CUST_PO_NO: 고객PO번호').font = Font(name='맑은 고딕', size=10, color='808080')

q8 = """
SELECT TOP 50 i.FACTORY_CODE, i.ORDER_NO, i.CONFIRM_DATE, i.ORDER_TYPE, i.ORDER_STATUS,
    i.MAT_CODE, i.LINE_CODE, i.ORD_QTY, i.SO_NO, i.SO_SEQ, i.CUSTOMER_CODE, i.CUST_PO_NO,
    i.ORD_START_TIME, i.ORD_END_TIME,
    o.ORD_STATUS AS MES_STATUS, o.ORD_OUT_QTY AS MES_OUT_QTY
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE = o.FACTORY_CODE AND i.ORDER_NO = o.ORDER_NO
WHERE i.CONFIRM_DATE >= '20260301'
ORDER BY i.CONFIRM_DATE DESC
"""
df8 = pd.read_sql(q8, conn)

r = 4
headers8 = ['공장','지시번호','확정일','유형','ERP상태','자재코드','라인','지시수량',
            'SO번호','SO순번','거래처','고객PO','시작일','종료일','MES상태','MES산출수량']
for i, h in enumerate(headers8, 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, len(headers8))

r += 1
for _, row in df8.iterrows():
    fname = factory_names.get(row['FACTORY_CODE'], row['FACTORY_CODE'])
    ws6.cell(row=r, column=1, value=fname)
    ws6.cell(row=r, column=2, value=row['ORDER_NO'])
    cd = str(row['CONFIRM_DATE'])
    ws6.cell(row=r, column=3, value=f'{cd[:4]}-{cd[4:6]}-{cd[6:]}' if len(cd)==8 else cd)
    ws6.cell(row=r, column=4, value=row['ORDER_TYPE'])
    ws6.cell(row=r, column=5, value=row['ORDER_STATUS'])
    ws6.cell(row=r, column=6, value=row['MAT_CODE'])
    ws6.cell(row=r, column=7, value=row['LINE_CODE'])
    ws6.cell(row=r, column=8, value=int(row['ORD_QTY']))
    ws6.cell(row=r, column=9, value=row['SO_NO'] or '')
    ws6.cell(row=r, column=10, value=int(row['SO_SEQ']) if row['SO_SEQ'] else '')
    ws6.cell(row=r, column=11, value=row['CUSTOMER_CODE'] or '')
    ws6.cell(row=r, column=12, value=row['CUST_PO_NO'] or '')
    st = str(row['ORD_START_TIME'] or '')
    ws6.cell(row=r, column=13, value=f'{st[:4]}-{st[4:6]}-{st[6:]}' if len(st)==8 else st)
    et = str(row['ORD_END_TIME'] or '')
    ws6.cell(row=r, column=14, value=f'{et[:4]}-{et[4:6]}-{et[6:]}' if len(et)==8 else et)
    ws6.cell(row=r, column=15, value=row['MES_STATUS'] or '')
    ws6.cell(row=r, column=16, value=int(row['MES_OUT_QTY']) if row['MES_OUT_QTY'] else 0)
    style_data(ws6, r, len(headers8))
    ws6.cell(row=r, column=8).number_format = '#,##0'
    ws6.cell(row=r, column=16).number_format = '#,##0'
    r += 1

auto_width(ws6)

# ============================================================
# Sheet 7: 코드값 사전
# ============================================================
ws7 = wb.create_sheet('7.코드값 사전')
ws7.sheet_properties.tabColor = '808080'

ws7.cell(row=1, column=1, value='MES 코드값 사전 (대시보드 참조용)').font = title_font

r = 3
ws7.cell(row=r, column=1, value='[1] 작업지시 상태 (ORD_STATUS)').font = section_font
r += 1
for i, h in enumerate(['코드','설명','의미'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 3)
r += 1
for code, desc, meaning in [
    ('CREATE','생성','ERP에서 지시 생성됨'),
    ('PLAN','계획','생산계획 수립됨'),
    ('WAIT','대기','작업 대기중 (라인 배정됨)'),
    ('CONFIRM','확정','작업 확정 (투입 전)'),
    ('PROCESS','진행','현재 생산 진행중'),
    ('CLOSE','마감','생산 완료 마감'),
    ('DELETE','삭제','취소된 지시')]:
    ws7.cell(row=r, column=1, value=code)
    ws7.cell(row=r, column=2, value=desc)
    ws7.cell(row=r, column=3, value=meaning)
    style_data(ws7, r, 3)
    r += 1

r += 1
ws7.cell(row=r, column=1, value='[2] 비가동 코드 (NONWORK_CODE)').font = section_font
r += 1
for i, h in enumerate(['분류','코드','설명'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 3)
r += 1
for code, desc in sorted(nwk_code_map.items()):
    cls = nwk_class_map.get(code[:2], code[:2])
    ws7.cell(row=r, column=1, value=cls)
    ws7.cell(row=r, column=2, value=code)
    ws7.cell(row=r, column=3, value=desc)
    style_data(ws7, r, 3)
    r += 1

r += 1
ws7.cell(row=r, column=1, value='[3] 라인유형 (LINE_TYPE)').font = section_font
r += 1
for i, h in enumerate(['코드','설명'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 2)
r += 1
for code, desc in [('BULK','제조(벌크)'),('FILLING','충진'),('TABLET','타정'),('PACKING','포장'),('BONDING','본딩')]:
    ws7.cell(row=r, column=1, value=code)
    ws7.cell(row=r, column=2, value=desc)
    style_data(ws7, r, 2)
    r += 1

r += 1
ws7.cell(row=r, column=1, value='[4] 트랜잭션 코드 (TRAN_CODE in MWIPLOTHIS)').font = section_font
r += 1
for i, h in enumerate(['코드','설명','용도'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 3)
r += 1
for code, desc, use in [
    ('CV','생산완료(Conversion)','LOT 완료 시 수량 기록, 핵심 생산 실적'),
    ('IS','입고(Issue/Store)','자재 투입'),
    ('RV','수량변경(Revision)','수량 보정'),
    ('SP','분할(Split)','LOT 분할'),
    ('MG','병합(Merge)','LOT 병합'),
    ('SC','폐기(Scrap)','불량/폐기 처리')]:
    ws7.cell(row=r, column=1, value=code)
    ws7.cell(row=r, column=2, value=desc)
    ws7.cell(row=r, column=3, value=use)
    style_data(ws7, r, 3)
    r += 1

r += 1
ws7.cell(row=r, column=1, value='[5] 공장코드').font = section_font
r += 1
for i, h in enumerate(['코드','공장명'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 2)
r += 1
for code, name in [('1100','퍼플카운티'),('1200','그린카운티'),('1300','3공장')]:
    ws7.cell(row=r, column=1, value=code)
    ws7.cell(row=r, column=2, value=name)
    style_data(ws7, r, 2)
    r += 1

auto_width(ws7)

# ============================================================
# Sheet 8: 대시보드 SQL 쿼리 모음
# ============================================================
ws8 = wb.create_sheet('8.SQL쿼리모음')
ws8.sheet_properties.tabColor = '375623'

ws8.cell(row=1, column=1, value='실시간 대시보드용 SQL 쿼리 모음').font = title_font
ws8.cell(row=2, column=1, value='아래 쿼리를 BI 도구나 Python에서 주기적으로 실행하여 대시보드를 구성합니다.').font = Font(name='맑은 고딕', size=10, color='808080')

queries = [
    ('Q1. 공장별 작업지시 상태 요약',
     """SELECT FACTORY_CODE,
  SUM(CASE WHEN ORD_STATUS='WAIT' THEN 1 ELSE 0 END) AS wait_cnt,
  SUM(CASE WHEN ORD_STATUS='WAIT' THEN ORD_QTY ELSE 0 END) AS wait_qty,
  SUM(CASE WHEN ORD_STATUS='CONFIRM' THEN 1 ELSE 0 END) AS cfm_cnt,
  SUM(CASE WHEN ORD_STATUS='CONFIRM' THEN ORD_QTY ELSE 0 END) AS cfm_qty,
  SUM(CASE WHEN ORD_STATUS='PROCESS' THEN 1 ELSE 0 END) AS proc_cnt,
  SUM(CASE WHEN ORD_STATUS='PROCESS' THEN ORD_QTY ELSE 0 END) AS proc_qty
FROM MWIPORDSTS WHERE ORD_STATUS NOT IN ('DELETE','CLOSE')
GROUP BY FACTORY_CODE""",
     '실시간 갱신 가능 (MWIPORDSTS는 상태 변경 시 즉시 반영)'),

    ('Q2. 오늘의 라인별 생산 실적',
     """SELECT h.FACTORY_CODE, h.LINE_CODE, l.LINE_DESC, l.LINE_TYPE,
  SUM(h.QTY) AS today_qty, COUNT(*) AS lot_cnt
FROM MWIPLOTHIS h
JOIN MWIPLINDEF l ON h.FACTORY_CODE=l.FACTORY_CODE AND h.LINE_CODE=l.LINE_CODE
WHERE h.TRAN_CODE='CV' AND CONVERT(VARCHAR, h.TRAN_TIME, 112) = CONVERT(VARCHAR, GETDATE(), 112)
GROUP BY h.FACTORY_CODE, h.LINE_CODE, l.LINE_DESC, l.LINE_TYPE
ORDER BY h.FACTORY_CODE, SUM(h.QTY) DESC""",
     '실시간 (CV 트랜잭션은 LOT 완료 즉시 기록)'),

    ('Q3. 현재 진행중인 비가동',
     """SELECT n.FACTORY_CODE, n.LINE_CODE, l.LINE_DESC,
  n.NONWORK_CODE, n.NONWORK_SECOND, n.PURE_NONWORK_SECOND,
  n.START_TIME_STR, n.END_TIME_STR, n.NONWORK_COMMENT
FROM MWIPNWKSTS n
LEFT JOIN MWIPLINDEF l ON n.FACTORY_CODE=l.FACTORY_CODE AND n.LINE_CODE=l.LINE_CODE
WHERE n.NONWORK_DATE = CONVERT(VARCHAR, GETDATE(), 112)
ORDER BY n.NONWORK_SECOND DESC""",
     '실시간 (비가동 발생 시 즉시 기록)'),

    ('Q4. 지시 대비 달성률 (당일)',
     """SELECT o.FACTORY_CODE, o.ORDER_NO, o.MAT_CODE, o.LINE_CODE,
  o.ORD_QTY, o.ORD_OUT_QTY,
  CASE WHEN o.ORD_QTY>0 THEN CAST(o.ORD_OUT_QTY*100.0/o.ORD_QTY AS DECIMAL(5,1)) ELSE 0 END AS achieve_pct
FROM MWIPORDSTS o
WHERE o.ORD_STATUS IN ('WAIT','PROCESS') AND o.PLAN_DATE = CONVERT(VARCHAR, GETDATE(), 112)
ORDER BY achieve_pct""",
     '실시간 (ORD_OUT_QTY는 LOT CV 시 갱신)'),

    ('Q5. SO(수주) 기준 생산 현황',
     """SELECT i.SO_NO, i.CUSTOMER_CODE, i.CUST_PO_NO,
  COUNT(*) AS order_cnt, SUM(i.ORD_QTY) AS total_ord_qty,
  SUM(ISNULL(o.ORD_OUT_QTY, 0)) AS total_out_qty,
  CASE WHEN SUM(i.ORD_QTY)>0
    THEN CAST(SUM(ISNULL(o.ORD_OUT_QTY,0))*100.0/SUM(i.ORD_QTY) AS DECIMAL(5,1))
    ELSE 0 END AS so_achieve_pct
FROM IWIPORDSTS i
LEFT JOIN MWIPORDSTS o ON i.FACTORY_CODE=o.FACTORY_CODE AND i.ORDER_NO=o.ORDER_NO
WHERE i.SO_NO IS NOT NULL AND i.SO_NO <> ''
GROUP BY i.SO_NO, i.CUSTOMER_CODE, i.CUST_PO_NO
ORDER BY so_achieve_pct""",
     'ERP 수주번호 기준 생산 진행 추적 (납기 관리 핵심)'),

    ('Q6. 라인별 가동률 (일별)',
     """SELECT n.FACTORY_CODE, n.LINE_CODE, n.NONWORK_DATE,
  SUM(n.NONWORK_SECOND) AS total_nwk_sec,
  SUM(n.PURE_NONWORK_SECOND) AS pure_nwk_sec,
  28800 AS std_work_sec,
  CAST((28800 - SUM(n.PURE_NONWORK_SECOND))*100.0/28800 AS DECIMAL(5,1)) AS operation_rate
FROM MWIPNWKSTS n
WHERE n.NONWORK_DATE >= CONVERT(VARCHAR, DATEADD(DAY,-7,GETDATE()), 112)
GROUP BY n.FACTORY_CODE, n.LINE_CODE, n.NONWORK_DATE
ORDER BY operation_rate""",
     '28800초(8시간) 기준 가동률, 최근 7일'),
]

r = 4
for title, sql, note in queries:
    ws8.cell(row=r, column=1, value=title).font = section_font
    r += 1
    ws8.cell(row=r, column=1, value='갱신주기:')
    ws8.cell(row=r, column=2, value=note)
    ws8.cell(row=r, column=2).font = Font(name='맑은 고딕', size=10, color='C55A11')
    r += 1
    ws8.cell(row=r, column=1, value=sql).font = Font(name='Consolas', size=9)
    ws8.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws8.row_dimensions[r].height = max(60, sql.count('\n') * 14)
    r += 2

ws8.column_dimensions['A'].width = 80
ws8.column_dimensions['B'].width = 50

# Save
out_path = 'C:/Users/user/Desktop/C&C/claude/MES_실시간_대시보드.xlsx'
wb.save(out_path)
conn.close()
print(f'Dashboard saved: {out_path}')
print(f'Sheets: {wb.sheetnames}')
