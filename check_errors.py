import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'
wb = openpyxl.load_workbook(file)

# 1. 수주_RAW P70~P75 #NAME? errors
print("=== 수주_RAW: P70~P75 #NAME? 수식 확인 ===")
ws = wb['수주_RAW']
for r in range(68, 78):
    c = ws.cell(r, 16)  # P column
    if c.value is not None:
        print(f"  P{r} = {c.value} (type={c.data_type})")

# 2. ④ 생산량 입력 row 17-18 #DIV/0!
print("\n=== ④ 생산량 입력: row 17~18 수식 확인 ===")
ws = wb['④ 생산량 입력']
for r in range(15, 21):
    vals = []
    for col in range(1, 32):
        c = ws.cell(r, col)
        if c.value is not None:
            prefix = '[F]' if c.data_type == 'f' else ''
            v = str(c.value)[:80]
            vals.append(f"{c.coordinate}={prefix}{v}")
    if vals:
        print(f"  Row {r}: {' | '.join(vals[:8])}")

# 3. ⑦ 생산Breakdown row 23-25 #DIV/0!
print("\n=== ⑦ 생산Breakdown: row 23~25 수식 확인 ===")
ws = wb['⑦ 생산Breakdown (OUTPUT)']
for r in range(21, 27):
    vals = []
    for col in range(1, 66):
        c = ws.cell(r, col)
        if c.value is not None and c.data_type == 'f':
            vals.append(f"{c.coordinate}={c.value[:80]}")
    if vals:
        print(f"  Row {r}: {' | '.join(vals[:5])}")

# Row labels
print("\n=== ⑦ 생산Breakdown: A열 레이블 ===")
for r in range(1, 32):
    c = ws.cell(r, 1)
    if c.value:
        print(f"  A{r} = {c.value}")

# 4. ④ 생산량 A열 labels
print("\n=== ④ 생산량 입력: A열 레이블 ===")
ws = wb['④ 생산량 입력']
for r in range(1, 25):
    c = ws.cell(r, 1)
    if c.value:
        print(f"  A{r} = {c.value}")

# 5. 생산량_RAW row count issue (1048576 = max rows)
print(f"\n=== 생산량_RAW: max_row = {wb['생산량_RAW'].max_row} (엑셀 최대행 = 1048576) ===")
ws = wb['생산량_RAW']
# Find actual last data row
last_data = 0
for r in range(ws.max_row, max(1, ws.max_row - 100), -1):
    c = ws.cell(r, 1)
    if c.value is not None:
        last_data = r
        break
print(f"  실제 마지막 데이터 행: {last_data}")
# Check row 2 to understand structure
row2 = []
for col in range(1, 25):
    c = ws.cell(2, col)
    if c.value is not None:
        prefix = '[F]' if c.data_type == 'f' else ''
        row2.append(f"{c.coordinate}={prefix}{str(c.value)[:50]}")
print(f"  Row 2: {' | '.join(row2)}")
