import pandas as pd
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file = r'C:\Users\user\Downloads\FY2602_결산입력_템플릿_260330.xlsx'

# Read 매출채권_RAW with pandas (faster than openpyxl for data analysis)
print("Reading 매출채권_RAW...")
df = pd.read_excel(file, sheet_name='매출채권_RAW', header=0)
print(f"Shape: {df.shape}")

# Find non-empty columns
non_empty = df.columns[df.notna().any() & (df != 0).any()].tolist()
print(f"\n실제 데이터가 있는 열: {len(non_empty)}개 (전체 {len(df.columns)}열)")
for c in non_empty:
    count = df[c].notna().sum()
    print(f"  {c}: {count}건")

# Also find which columns ⑪ references
import openpyxl, re
wb = openpyxl.load_workbook(file)
ws_out = wb['⑪ 매출채권 (OUTPUT)']
referenced_cols = set()
for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row, max_col=ws_out.max_column):
    for c in row:
        if c.data_type == 'f' and c.value and '매출채권_RAW' in str(c.value):
            refs = re.findall(r"매출채권_RAW!\$([A-Z]+)", str(c.value))
            referenced_cols.update(refs)
print(f"\n⑪에서 참조하는 매출채권_RAW 열: {sorted(referenced_cols)}")
